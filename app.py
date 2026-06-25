import cgi
import base64
import hashlib
import hmac
import html
import io
import json
import math
import mimetypes
import operator
import os
import re
import subprocess
import sys
import time
import uuid
import tempfile
import zipfile
import smtplib
import ssl
from dataclasses import dataclass, asdict
from datetime import datetime
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from http.cookies import SimpleCookie
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from urllib.parse import parse_qs, quote, urlparse
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
DATA_DIR = ROOT / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
KB_FILE = DATA_DIR / "knowledge_base.json"
AUTH_FILE = DATA_DIR / "auth.json"
AUTH_XLSX = DATA_DIR / "auth.xlsx"
PENDING_AUTH_FILE = DATA_DIR / "pending_auth.json"
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
ALLOWED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
DEFAULT_OPENAI_MODEL = "gpt-4o-mini"
DEFAULT_OPENAI_IMAGE_MODEL = "gpt-image-2"
DEFAULT_APP_LOGIN_USERNAME = "admin"
DEFAULT_APP_LOGIN_PASSWORD = "admin123"
DEFAULT_APP_SESSION_SECRET = "jaffagpt-session-secret"
DEFAULT_PASSWORD_ITERATIONS = 120_000
DEFAULT_OTP_ITERATIONS = 10_000
MAX_SOURCE_CHARS = 1800
STOP_WORDS = {
    "the",
    "and",
    "for",
    "with",
    "this",
    "that",
    "from",
    "find",
    "name",
    "pdf",
    "document",
    "documents",
    "uploaded",
    "upload",
    "file",
    "files",
    "what",
    "when",
    "where",
    "which",
    "who",
    "why",
    "how",
    "tell",
    "give",
    "show",
    "answer",
    "question",
    "please",
    "about",
    "into",
    "there",
    "their",
    "your",
    "you",
    "are",
    "was",
    "were",
    "have",
    "has",
    "had",
    "can",
    "could",
    "should",
}
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", DEFAULT_OPENAI_MODEL).strip() or DEFAULT_OPENAI_MODEL
OPENAI_IMAGE_MODEL = (
    os.environ.get("OPENAI_IMAGE_MODEL", DEFAULT_OPENAI_IMAGE_MODEL).strip()
    or DEFAULT_OPENAI_IMAGE_MODEL
)
OCR_SPACE_API_KEY = os.environ.get("OCR_SPACE_API_KEY", "").strip()
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "").strip()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()
ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-20250514").strip() or "claude-sonnet-4-20250514"
APP_LOGIN_USERNAME = os.environ.get("APP_LOGIN_USERNAME", DEFAULT_APP_LOGIN_USERNAME).strip() or DEFAULT_APP_LOGIN_USERNAME
APP_LOGIN_PASSWORD = os.environ.get("APP_LOGIN_PASSWORD", DEFAULT_APP_LOGIN_PASSWORD).strip() or DEFAULT_APP_LOGIN_PASSWORD
APP_SESSION_SECRET = os.environ.get("APP_SESSION_SECRET", DEFAULT_APP_SESSION_SECRET).strip() or DEFAULT_APP_SESSION_SECRET
SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587") or 587)
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "").strip()
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "").strip()
SMTP_FROM_EMAIL = os.environ.get("SMTP_FROM_EMAIL", "").strip()
SMTP_USE_TLS = (os.environ.get("SMTP_USE_TLS", "true").strip().lower() not in {"0", "false", "no"})

DEFAULT_GEMINI_MODEL = "gemini-1.5-flash"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", DEFAULT_GEMINI_MODEL).strip() or DEFAULT_GEMINI_MODEL



class TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        text = data.strip()
        if text:
            self.parts.append(text)

    def text(self) -> str:
        return "\n".join(self.parts)
RETRIABLE_OPENAI_STATUS = {408, 409, 429, 500, 502, 503, 504}
PDF_MISSING_TEXT = (
    "PDF text extraction needs PyMuPDF installed. "
    "Install it with: pip install pymupdf"
)
TRANSLATION_HINTS = {
    "Hindi": {
        "hello": "namaste",
        "help": "madad",
        "document": "dastavez",
        "deadline": "antim tithi",
        "fee": "shulk",
        "question": "sawal",
        "answer": "jawab",
    },
    "Telugu": {
        "hello": "namaskaram",
        "help": "sahayam",
        "document": "patram",
        "deadline": "chivari tariku",
        "fee": "rusumu",
        "question": "prasna",
        "answer": "samadhanam",
    },
    "Tamil": {
        "hello": "vanakkam",
        "help": "uthavi",
        "document": "aavanam",
        "deadline": "kadaisi naal",
        "fee": "kattanam",
        "question": "kelvi",
        "answer": "badil",
    },
    "Kannada": {
        "hello": "namaskara",
        "help": "sahaya",
        "document": "dakhale",
        "deadline": "koneya dina",
        "fee": "shulka",
        "question": "prashne",
        "answer": "uttara",
    },
}

try:
    import openai_config

    config_key = getattr(openai_config, "OPENAI_API_KEY", "")
    config_model = getattr(openai_config, "OPENAI_MODEL", DEFAULT_OPENAI_MODEL)
    config_image_model = getattr(openai_config, "OPENAI_IMAGE_MODEL", DEFAULT_OPENAI_IMAGE_MODEL)
    config_ocr_key = getattr(openai_config, "OCR_SPACE_API_KEY", "")
    config_google_client_id = getattr(openai_config, "GOOGLE_CLIENT_ID", "")
    config_anthropic_key = getattr(openai_config, "ANTHROPIC_API_KEY", "")
    config_anthropic_model = getattr(openai_config, "ANTHROPIC_MODEL", "claude-sonnet-4-20250514")
    OPENAI_API_KEY = OPENAI_API_KEY or str(config_key).strip()
    OPENAI_MODEL = (os.environ.get("OPENAI_MODEL") or str(config_model)).strip() or DEFAULT_OPENAI_MODEL
    OPENAI_IMAGE_MODEL = (
        os.environ.get("OPENAI_IMAGE_MODEL") or str(config_image_model)
    ).strip() or DEFAULT_OPENAI_IMAGE_MODEL
    OCR_SPACE_API_KEY = OCR_SPACE_API_KEY or str(config_ocr_key).strip()
    GOOGLE_CLIENT_ID = GOOGLE_CLIENT_ID or str(config_google_client_id).strip()
    ANTHROPIC_API_KEY = ANTHROPIC_API_KEY or str(config_anthropic_key).strip()
    ANTHROPIC_MODEL = (os.environ.get("ANTHROPIC_MODEL") or str(config_anthropic_model)).strip() or "claude-sonnet-4-20250514"

    config_gemini_key = getattr(openai_config, "GEMINI_API_KEY", "")
    config_gemini_model = getattr(openai_config, "GEMINI_MODEL", DEFAULT_GEMINI_MODEL)
    GEMINI_API_KEY = GEMINI_API_KEY or str(config_gemini_key).strip()
    GEMINI_MODEL = (os.environ.get("GEMINI_MODEL") or str(config_gemini_model)).strip() or DEFAULT_GEMINI_MODEL

    config_login_username = getattr(openai_config, "APP_LOGIN_USERNAME", "")
    config_login_password = getattr(openai_config, "APP_LOGIN_PASSWORD", "")
    config_session_secret = getattr(openai_config, "APP_SESSION_SECRET", "")
    APP_LOGIN_USERNAME = str(config_login_username).strip() or APP_LOGIN_USERNAME
    APP_LOGIN_PASSWORD = str(config_login_password).strip() or APP_LOGIN_PASSWORD
    APP_SESSION_SECRET = str(config_session_secret).strip() or APP_SESSION_SECRET

    config_smtp_host = getattr(openai_config, "SMTP_HOST", "")
    config_smtp_port = getattr(openai_config, "SMTP_PORT", SMTP_PORT)
    config_smtp_username = getattr(openai_config, "SMTP_USERNAME", "")
    config_smtp_password = getattr(openai_config, "SMTP_PASSWORD", "")
    config_smtp_from_email = getattr(openai_config, "SMTP_FROM_EMAIL", "")
    config_smtp_use_tls = getattr(openai_config, "SMTP_USE_TLS", SMTP_USE_TLS)
    SMTP_HOST = str(config_smtp_host).strip() or SMTP_HOST
    try:
        SMTP_PORT = int(config_smtp_port)
    except (TypeError, ValueError):
        pass
    SMTP_USERNAME = str(config_smtp_username).strip() or SMTP_USERNAME
    SMTP_PASSWORD = str(config_smtp_password).strip() or SMTP_PASSWORD
    SMTP_FROM_EMAIL = str(config_smtp_from_email).strip() or SMTP_FROM_EMAIL
    SMTP_USE_TLS = bool(config_smtp_use_tls)
except ImportError:
    pass



@dataclass
class Document:
    id: str
    name: str
    category: str
    text: str
    created_at: float


def ensure_dirs() -> None:
    STATIC_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    if not KB_FILE.exists():
        KB_FILE.write_text("[]", encoding="utf-8")
    if not AUTH_XLSX.exists():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(["username", "email", "salt", "password_hash", "created_at"])
        workbook.save(AUTH_XLSX)
    if not PENDING_AUTH_FILE.exists():
        PENDING_AUTH_FILE.write_text("[]", encoding="utf-8")


def load_documents() -> list[Document]:
    ensure_dirs()
    try:
        raw_docs = json.loads(KB_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        raw_docs = []
    documents = []
    for doc in raw_docs:
        try:
            documents.append(Document(**doc))
        except TypeError:
            continue
    return documents


def save_documents(documents: list[Document]) -> None:
    ensure_dirs()
    KB_FILE.write_text(
        json.dumps([asdict(doc) for doc in documents], indent=2),
        encoding="utf-8",
    )


def otp_digest(email: str, otp: str, token: str | None = None) -> tuple[str, str]:
    token = token or str(uuid.uuid4())
    secret = f"{APP_SESSION_SECRET}:{email}:{token}".encode("utf-8")
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        otp.encode("utf-8"),
        secret,
        DEFAULT_OTP_ITERATIONS,
    )
    return token, base64.b64encode(digest).decode("ascii")


def verify_otp(email: str, otp: str, token: str, digest_b64: str) -> bool:
    _, check_digest = otp_digest(email, otp, token)
    return hmac.compare_digest(check_digest, digest_b64)


def hash_password(password: str, salt: bytes | None = None) -> tuple[str, str]:
    salt = salt or os.urandom(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        DEFAULT_PASSWORD_ITERATIONS,
    )
    return base64.b64encode(salt).decode("ascii"), base64.b64encode(digest).decode("ascii")


def verify_password(password: str, salt_b64: str, hash_b64: str) -> bool:
    try:
        salt = base64.b64decode(salt_b64.encode("ascii"))
    except (ValueError, UnicodeEncodeError):
        return False
    _, check_hash = hash_password(password, salt)
    return hmac.compare_digest(check_hash, hash_b64)


def normalize_email(email: str) -> str:
    return email.strip().lower()


def validate_email(email: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", email.strip()))


def normalize_contact(contact: str) -> str:
    contact = str(contact or "").strip()
    if not contact:
        return ""
    if "@" in contact:
        return normalize_email(contact)
    return re.sub(r"[\s().-]+", "", contact)


def validate_contact(contact: str) -> bool:
    contact = str(contact or "").strip()
    if not contact:
        return False
    if validate_email(contact):
        return True
    return bool(re.fullmatch(r"\+?[0-9][0-9\s().-]{6,}", contact))


def load_users() -> list[dict[str, str]]:
    ensure_dirs()
    legacy_users: list[dict[str, str]] = []
    if AUTH_FILE.exists():
        try:
            parsed_legacy = json.loads(AUTH_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            parsed_legacy = []
        if isinstance(parsed_legacy, list):
            legacy_users = [user for user in parsed_legacy if isinstance(user, dict) and user.get("username")]

    users: list[dict[str, str]] = []
    try:
        workbook = load_workbook(AUTH_XLSX)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        rows = []

    if rows:
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        for row in rows[1:]:
            entry = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
            username = str(entry.get("username", "")).strip()
            email = str(entry.get("email", "")).strip()
            salt = str(entry.get("salt", "")).strip()
            password_hash = str(entry.get("password_hash", "")).strip()
            provider = str(entry.get("provider", "password") or "password").strip() or "password"
            created_at = entry.get("created_at", time.time())
            if username and salt and password_hash:
                try:
                    created_at_value = float(created_at)
                except (TypeError, ValueError):
                    created_at_value = time.time()
                users.append(
                    {
                        "username": username,
                        "email": email,
                        "salt": salt,
                        "password_hash": password_hash,
                        "provider": provider,
                        "created_at": created_at_value,
                    }
                )

    if legacy_users:
        merged: dict[str, dict[str, str]] = {user["username"]: user for user in users if user.get("username")}
        for user in legacy_users:
            username = str(user.get("username", "")).strip()
            email = str(user.get("email", "")).strip()
            salt = str(user.get("salt", "")).strip()
            password_hash = str(user.get("password_hash", "")).strip()
            created_at = user.get("created_at", time.time())
            if username and salt and password_hash and username not in merged:
                merged[username] = {
                    "username": username,
                    "email": email,
                    "salt": salt,
                    "password_hash": password_hash,
                    "provider": str(user.get("provider", "password") or "password"),
                    "created_at": created_at,
                }
        users = list(merged.values())
        try:
            save_users(users)
        except PermissionError:
            pass
        try:
            AUTH_FILE.unlink()
        except OSError:
            pass

    if not users:
        salt, password_hash = hash_password(DEFAULT_APP_LOGIN_PASSWORD)
        users = [
            {
                "username": DEFAULT_APP_LOGIN_USERNAME,
                "email": "",
                "salt": salt,
                "password_hash": password_hash,
                    "provider": "password",
                "created_at": time.time(),
            }
        ]
        try:
            save_users(users)
        except PermissionError:
            pass
    return users


def save_users(users: list[dict[str, str]]) -> None:
    ensure_dirs()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"
    sheet.append(["username", "email", "salt", "password_hash", "provider", "created_at"])
    for user in users:
        if not isinstance(user, dict):
            continue
        username = str(user.get("username", "")).strip()
        email = str(user.get("email", "")).strip()
        salt = str(user.get("salt", "")).strip()
        password_hash = str(user.get("password_hash", "")).strip()
        provider = str(user.get("provider", "password") or "password").strip() or "password"
        created_at = user.get("created_at", time.time())
        if username and salt and password_hash:
            sheet.append([username, email, salt, password_hash, provider, created_at])
    workbook.save(AUTH_XLSX)


def find_user(username: str) -> dict[str, str] | None:
    username = username.strip()
    for user in load_users():
        if user.get("username") == username:
            return user
    return None


def find_user_by_contact(contact: str) -> dict[str, str] | None:
    contact = normalize_contact(contact)
    for user in load_users():
        if normalize_contact(str(user.get("email", ""))) == contact and contact:
            return user
    return None


def find_user_by_email(email: str) -> dict[str, str] | None:
    return find_user_by_contact(email)


def create_user(username: str, email: str, password: str) -> tuple[bool, str]:
    username = username.strip()
    email = str(email).strip()
    password = password.strip()
    if not re.fullmatch(r"[A-Za-z0-9_. -]{3,32}", username):
        return False, "Name must be 3-32 characters and may include letters, numbers, spaces, dot, dash, or underscore."
    if not validate_contact(email):
        return False, "Enter a valid email address or phone number."
    if len(password) < 6:
        return False, "Password must be at least 6 characters long."
    if find_user(username):
        return False, "That name already exists."
    if find_user_by_contact(email):
        return False, "That email or number is already registered."

    salt, password_hash = hash_password(password)
    users = load_users()
    users.append(
        {
            "username": username,
            "email": email,
            "salt": salt,
            "password_hash": password_hash,
            "provider": "password",
            "created_at": time.time(),
        }
    )
    save_users(users)
    return True, "Account created."


def update_user_password(contact: str, password: str) -> tuple[bool, str]:
    contact = normalize_contact(contact)
    if not validate_contact(contact):
        return False, "Enter a valid email address or phone number."
    if len(password.strip()) < 6:
        return False, "Password must be at least 6 characters long."

    users = load_users()
    target_index = None
    for index, user in enumerate(users):
        if normalize_contact(str(user.get("email", ""))) == contact or normalize_contact(str(user.get("username", ""))) == contact:
            target_index = index
            break

    if target_index is None:
        return False, "No account was found for that email or number."

    salt, password_hash = hash_password(password)
    users[target_index]["salt"] = salt
    users[target_index]["password_hash"] = password_hash
    save_users(users)
    return True, "Password updated."


def unique_username(base_username: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(base_username).strip())
    cleaned = re.sub(r"[^A-Za-z0-9_. -]", "", cleaned).strip(" .-_")
    if not cleaned:
        cleaned = "Google User"

    candidate = cleaned[:32]
    suffix = 2
    while find_user(candidate):
        suffix_text = f" {suffix}"
        candidate = f"{cleaned[: max(1, 32 - len(suffix_text))]}{suffix_text}"
        suffix += 1
    return candidate


def verify_google_id_token(credential: str) -> tuple[bool, dict[str, str] | str]:
    if not GOOGLE_CLIENT_ID:
        return False, "Google sign-in is not configured on this server yet."

    token_url = f"https://oauth2.googleapis.com/tokeninfo?id_token={quote(credential)}"
    request = Request(token_url, headers={"User-Agent": "JaffaGPT/1.0"})
    try:
        with urlopen(request, timeout=15) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        try:
            error_payload = json.loads(exc.read().decode("utf-8"))
            message = error_payload.get("error_description") or error_payload.get("error") or "Google sign-in failed."
        except Exception:
            message = "Google sign-in failed."
        return False, str(message)
    except (URLError, json.JSONDecodeError, OSError):
        return False, "Google sign-in could not be verified right now."

    if str(payload.get("aud", "")).strip() != GOOGLE_CLIENT_ID:
        return False, "Google sign-in token was issued for a different client."

    issuer = str(payload.get("iss", "")).strip()
    if issuer not in {"accounts.google.com", "https://accounts.google.com"}:
        return False, "Google sign-in token issuer was not accepted."

    if str(payload.get("email_verified", "")).lower() != "true":
        return False, "Google account email is not verified."

    email = str(payload.get("email", "")).strip()
    if not email:
        return False, "Google account email is missing."

    return True, payload


def load_pending_registrations() -> list[dict[str, str]]:
    ensure_dirs()
    try:
        pending = json.loads(PENDING_AUTH_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        pending = []
    return [item for item in pending if isinstance(item, dict) and item.get("pending_id")]


def save_pending_registrations(pending: list[dict[str, str]]) -> None:
    ensure_dirs()
    PENDING_AUTH_FILE.write_text(json.dumps(pending, indent=2), encoding="utf-8")


def add_pending_registration(username: str, email: str, password: str, otp: str) -> dict[str, str]:
    salt, password_hash = hash_password(password)
    token, otp_hash = otp_digest(email, otp)
    pending = load_pending_registrations()
    pending = [item for item in pending if item.get("email") != email and item.get("username") != username]
    record = {
        "pending_id": token,
        "action": "register",
        "username": username,
        "contact": normalize_contact(email),
        "email": normalize_contact(email),
        "salt": salt,
        "password_hash": password_hash,
        "otp_hash": otp_hash,
        "created_at": time.time(),
        "expires_at": time.time() + 600,
    }
    pending.append(record)
    save_pending_registrations(pending)
    return record


def find_pending_registration(pending_id: str) -> dict[str, str] | None:
    for item in load_pending_registrations():
        if item.get("pending_id") == pending_id and item.get("action", "register") == "register":
            return item
    return None


def add_pending_password_reset(contact: str, password: str, otp: str) -> dict[str, str]:
    salt, password_hash = hash_password(password)
    token, otp_hash = otp_digest(contact, otp)
    pending = load_pending_registrations()
    pending = [item for item in pending if item.get("email") != contact and item.get("pending_id") != token]
    record = {
        "pending_id": token,
        "action": "reset",
        "username": "",
        "contact": normalize_contact(contact),
        "email": normalize_contact(contact),
        "salt": salt,
        "password_hash": password_hash,
        "otp_hash": otp_hash,
        "created_at": time.time(),
        "expires_at": time.time() + 600,
    }
    pending.append(record)
    save_pending_registrations(pending)
    return record


def find_pending_password_reset(pending_id: str) -> dict[str, str] | None:
    for item in load_pending_registrations():
        if item.get("pending_id") == pending_id and item.get("action") == "reset":
            return item
    return None


def remove_pending_registration(pending_id: str) -> None:
    pending = [item for item in load_pending_registrations() if item.get("pending_id") != pending_id]
    save_pending_registrations(pending)


def send_otp_code(contact: str, otp: str, username: str, purpose: str = "verification") -> tuple[bool, str, str | None]:
    contact = str(contact or "").strip()
    if validate_email(contact):
        from_email = SMTP_FROM_EMAIL or SMTP_USERNAME
        if not SMTP_HOST or not from_email:
            return True, "Email delivery is not configured here, so the OTP will be shown on screen.", otp

        message = EmailMessage()
        subject = "Your JaffaGPT verification code" if purpose == "verification" else "Your JaffaGPT password reset code"
        message["Subject"] = subject
        message["From"] = from_email
        message["To"] = contact
        message.set_content(
            f"Hello {username},\n\n"
            f"Your JaffaGPT code is: {otp}\n\n"
            "This code expires in 10 minutes.\n"
            "If you did not request this action, you can ignore this email."
        )

        try:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
                if SMTP_USE_TLS:
                    server.starttls(context=ssl.create_default_context())
                if SMTP_USERNAME:
                    server.login(SMTP_USERNAME, SMTP_PASSWORD)
                server.send_message(message)
            return True, "OTP sent.", None
        except Exception as exc:
            return False, f"Could not send verification email: {exc}", None

    if re.fullmatch(r"\+?[0-9][0-9\s().-]{6,}", contact):
        return True, "SMS is not configured here, so the OTP will be shown on screen.", otp

    return False, "Enter a valid email address or phone number.", None


def finalize_pending_registration(pending_id: str, otp: str) -> tuple[bool, str, str | None]:
    pending = find_pending_registration(pending_id)
    if not pending:
        return False, "Verification session expired. Please create the account again.", None
    if float(pending.get("expires_at", 0)) < time.time():
        remove_pending_registration(pending_id)
        return False, "Verification code expired. Please create the account again.", None
    contact = str(pending.get("contact", pending.get("email", ""))).strip()
    if not verify_otp(contact, otp.strip(), str(pending.get("pending_id", "")), str(pending.get("otp_hash", ""))):
        return False, "Invalid verification code.", None

    username = str(pending.get("username", "")).strip()
    email = contact
    salt = str(pending.get("salt", "")).strip()
    password_hash = str(pending.get("password_hash", "")).strip()
    if not username or not salt or not password_hash:
        return False, "Verification record is incomplete.", None

    users = load_users()
    if find_user(username):
        remove_pending_registration(pending_id)
        return False, "That username already exists.", None

    users.append(
        {
            "username": username,
            "email": email,
            "salt": salt,
            "password_hash": password_hash,
            "created_at": time.time(),
        }
    )
    try:
        save_users(users)
    except PermissionError:
        return False, "The login spreadsheet is open in Excel. Close it and try again.", None
    remove_pending_registration(pending_id)
    return True, "Account verified.", username


def finalize_pending_password_reset(pending_id: str, otp: str) -> tuple[bool, str, str | None]:
    pending = find_pending_password_reset(pending_id)
    if not pending:
        return False, "Reset session expired. Please request a new code.", None
    if float(pending.get("expires_at", 0)) < time.time():
        remove_pending_registration(pending_id)
        return False, "Reset code expired. Please request a new code.", None

    contact = str(pending.get("contact", pending.get("email", ""))).strip()
    if not verify_otp(contact, otp.strip(), str(pending.get("pending_id", "")), str(pending.get("otp_hash", ""))):
        return False, "Invalid reset code.", None

    new_password_hash = str(pending.get("password_hash", "")).strip()
    new_salt = str(pending.get("salt", "")).strip()
    users = load_users()
    updated = False
    for user in users:
        if normalize_contact(str(user.get("email", ""))) == normalize_contact(contact) or normalize_contact(str(user.get("username", ""))) == normalize_contact(contact):
            user["salt"] = new_salt
            user["password_hash"] = new_password_hash
            updated = True
            break

    if not updated:
        remove_pending_registration(pending_id)
        return False, "No account was found for that email or number.", None

    try:
        save_users(users)
    except PermissionError:
        return False, "The login spreadsheet is open in Excel. Close it and try again.", None

    remove_pending_registration(pending_id)
    return True, "Password updated.", contact


def add_document(name: str, category: str, text: str, doc_id: str | None = None) -> Document:
    doc = Document(
        id=doc_id or str(uuid.uuid4()),
        name=name,
        category=category,
        text=text,
        created_at=time.time(),
    )
    documents = load_documents()
    documents.append(doc)
    save_documents(documents)
    return doc


def decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def looks_like_binary(content: bytes) -> bool:
    sample = content[:4096]
    return b"\x00" in sample


def normalize_text(text: str) -> str:
    text = html.unescape(text)
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_html_text(content: bytes) -> str:
    parser = TextHTMLParser()
    parser.feed(decode_text(content))
    return normalize_text(parser.text())


def strip_rtf(content: bytes) -> str:
    text = decode_text(content)
    text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text)
    text = re.sub(r"\\[a-zA-Z]+\d* ?", " ", text)
    text = re.sub(r"[{}]", " ", text)
    return normalize_text(text)


def extract_office_text(content: bytes, suffix: str) -> str:
    targets = {
        ".docx": ("word/document.xml",),
        ".pptx": ("ppt/slides/slide",),
        ".xlsx": ("xl/sharedStrings.xml", "xl/worksheets/sheet"),
    }.get(suffix, ())
    parts = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            for name in archive.namelist():
                if not name.endswith(".xml"):
                    continue
                if targets and not any(name.startswith(target) for target in targets):
                    continue
                xml_text = archive.read(name).decode("utf-8", errors="ignore")
                xml_text = re.sub(r"<[^>]+>", " ", xml_text)
                clean = normalize_text(xml_text)
                if clean:
                    parts.append(clean)
    except zipfile.BadZipFile as exc:
        raise ValueError(f"Could not read text from this {suffix.upper()} file.") from exc
    return normalize_text("\n".join(parts))


def extract_email_text(content: bytes) -> str:
    message = BytesParser(policy=policy.default).parsebytes(content)
    parts = []
    if message.get("subject"):
        parts.append(f"Subject: {message.get('subject')}")
    for part in message.walk():
        if part.get_content_maintype() == "multipart":
            continue
        if part.get_content_type() in {"text/plain", "text/html"}:
            payload = part.get_content()
            if part.get_content_type() == "text/html":
                parser = TextHTMLParser()
                parser.feed(str(payload))
                payload = parser.text()
            parts.append(str(payload))
    return normalize_text("\n".join(parts))


OCR_LANG_MAP = {
    "English": {"space": "eng", "tess": "eng"},
    "English (US)": {"space": "eng", "tess": "eng"},
    "English (UK)": {"space": "eng", "tess": "eng"},
    "Hindi": {"space": "hin", "tess": "hin"},
    "Telugu": {"space": "tel", "tess": "tel"},
    "Tamil": {"space": "tam", "tess": "tam"},
    "Kannada": {"space": "kan", "tess": "kan"},
    "Malayalam": {"space": "mal", "tess": "mal"},
    "Marathi": {"space": "mar", "tess": "mar"},
    "Bengali": {"space": "ben", "tess": "ben"},
    "Gujarati": {"space": "guj", "tess": "guj"},
    "Punjabi": {"space": "pan", "tess": "pan"},
    "Urdu": {"space": "urd", "tess": "urd"},
    "Arabic": {"space": "ara", "tess": "ara"},
    "Chinese": {"space": "chs", "tess": "chi_sim"},
    "Chinese (Taiwan)": {"space": "cht", "tess": "chi_tra"},
    "Japanese": {"space": "jpn", "tess": "jpn"},
    "Korean": {"space": "kor", "tess": "kor"},
    "French": {"space": "fre", "tess": "fra"},
    "Spanish": {"space": "spa", "tess": "spa"},
    "Spanish (Mexico)": {"space": "spa", "tess": "spa"},
    "German": {"space": "ger", "tess": "deu"},
    "Italian": {"space": "ita", "tess": "ita"},
    "Portuguese": {"space": "por", "tess": "por"},
    "Portuguese (Portugal)": {"space": "por", "tess": "por"},
    "Russian": {"space": "rus", "tess": "rus"},
    "Dutch": {"space": "dut", "tess": "nld"},
    "Turkish": {"space": "tur", "tess": "tur"},
    "Vietnamese": {"space": "vie", "tess": "vie"},
    "Swedish": {"space": "swe", "tess": "swe"},
    "Norwegian": {"space": "nor", "tess": "nor"},
    "Danish": {"space": "dan", "tess": "dan"},
    "Finnish": {"space": "fin", "tess": "fin"},
    "Czech": {"space": "cze", "tess": "ces"},
    "Hungarian": {"space": "hun", "tess": "hun"},
    "Polish": {"space": "pol", "tess": "pol"},
    "Romanian": {"space": "rum", "tess": "ron"},
    "Greek": {"space": "gre", "tess": "ell"},
    "Ukrainian": {"space": "ukr", "tess": "ukr"},
}


def local_or_free_ocr(content: bytes, language: str) -> str:
    lang_info = OCR_LANG_MAP.get(language, {"space": "eng", "tess": "eng"})
    
    # 1. Try pytesseract first
    try:
        import pytesseract
        from PIL import Image
        
        tess_cmd = os.environ.get("TESSERACT_CMD", "")
        if tess_cmd:
            pytesseract.pytesseract.tesseract_cmd = tess_cmd
            
        img = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(img, lang=lang_info["tess"])
        if text.strip():
            return text.strip()
    except Exception:
        pass
        
    # 2. Try EasyOCR
    try:
        import easyocr
        easyocr_lang = lang_info["tess"][:2]
        langs = [easyocr_lang]
        if easyocr_lang == "en":
            langs = ["en"]
        elif easyocr_lang == "ch":
            langs = ["ch_sim"] if lang_info["tess"] == "chi_sim" else ["ch_tra"]
            
        reader = easyocr.Reader(langs, gpu=False)
        results = reader.readtext(content)
        text = "\n".join(res[1] for res in results)
        if text.strip():
            return text.strip()
    except Exception:
        pass

    # 3. Fallback to free OCR.space API
    try:
        api_key = OCR_SPACE_API_KEY.strip()
        if not api_key:
            api_key = "helloworld"
            
        b64_data = base64.b64encode(content).decode("utf-8")
        payload = {
            "apikey": api_key,
            "base64Image": f"data:image/png;base64,{b64_data}",
            "language": lang_info["space"],
            "OCREngine": "2",
        }
        
        from urllib.parse import urlencode
        body = urlencode(payload).encode("utf-8")
        request = Request(
            "https://api.ocr.space/parse/image",
            data=body,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
            },
            method="POST"
        )
        
        with urlopen(request, timeout=30) as response:
            res_data = json.loads(response.read().decode("utf-8"))
            
        if res_data.get("IsErroredOnProcessing"):
            err_msg = res_data.get("ErrorMessage")
            if isinstance(err_msg, list) and err_msg:
                err_msg = err_msg[0]
            raise ValueError(f"OCR.space error: {err_msg}")
            
        parsed_results = res_data.get("ParsedResults", [])
        text_parts = []
        for result in parsed_results:
            text_parts.append(result.get("ParsedText", ""))
            
        text = "\n".join(text_parts).strip()
        if text:
            return text
            
    except Exception as exc:
        raise RuntimeError(f"OCR failed. Local engines unavailable and API fallback failed: {exc}")

    raise RuntimeError("OCR failed to extract any text from the image.")


def extract_text(file_name: str, content: bytes, language: str = "English") -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix in {
        ".txt",
        ".md",
        ".csv",
        ".tsv",
        ".log",
        ".ini",
        ".cfg",
        ".conf",
        ".py",
        ".js",
        ".ts",
        ".css",
        ".java",
        ".c",
        ".cpp",
        ".h",
        ".json",
        ".xml",
        ".yaml",
        ".yml",
    }:
        return normalize_text(decode_text(content))

    if suffix == ".pdf":
        try:
            import fitz  # type: ignore
        except ImportError as exc:
            raise RuntimeError(PDF_MISSING_TEXT) from exc

        try:
            with fitz.open(stream=content, filetype="pdf") as pdf:
                text = "\n".join(page.get_text() for page in pdf).strip()
        except Exception as exc:
            raise ValueError("Could not read text from this PDF.") from exc

        if not text:
            ocr_text_parts = []
            try:
                with fitz.open(stream=content, filetype="pdf") as pdf:
                    for page in pdf:
                        pix = page.get_pixmap()
                        png_bytes = pix.tobytes("png")
                        page_text = local_or_free_ocr(png_bytes, language)
                        if page_text:
                            ocr_text_parts.append(page_text)
                text = "\n\n".join(ocr_text_parts).strip()
            except Exception as exc:
                raise ValueError(
                    f"No selectable text was found in this PDF. Attempted OCR but encountered an error: {exc}"
                )

            if not text:
                raise ValueError(
                    "No selectable text or OCR-readable text was found in this PDF."
                )
        return normalize_text(text)


    if suffix in {".html", ".htm"}:
        return extract_html_text(content)

    if suffix == ".rtf":
        return strip_rtf(content)

    if suffix in {".docx", ".pptx", ".xlsx"}:
        text = extract_office_text(content, suffix)
        if not text:
            raise ValueError(f"No readable text was found in this {suffix.upper()} file.")
        return text

    if suffix in {".eml", ".msg"}:
        text = extract_email_text(content)
        if text:
            return text

    if not looks_like_binary(content):
        text = normalize_text(decode_text(content))
        if text:
            return text

    return (
        f"{file_name} was uploaded and kept in the local document store, but this file "
        "type does not expose readable text with the built-in extractor. Convert it to "
        "PDF, TXT, DOCX, HTML, CSV, or add OCR/extraction support to ask detailed "
        "questions about its contents."
    )


def remove_uploaded_files(doc_id: str) -> None:
    for path in UPLOAD_DIR.glob(f"{doc_id}-*"):
        if path.is_file():
            path.unlink()


def tokenize(text: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-zA-Z0-9]{3,}", text.lower())
        if token
        not in {
            "the",
            "and",
            "for",
            "with",
            "this",
            "that",
            "from",
            "find",
            "name",
            "pdf",
            "document",
            "uploaded",
        }
    }


def is_document_lookup(question: str) -> bool:
    text = question.lower()
    return bool(
        re.search(
            r"\b(pdf|document|file|uploaded|source|resume|cv|ats|score|certificate|form|name)\b",
            text,
        )
    )


def is_ats_score_request(question: str) -> bool:
    text = question.lower()
    return bool(
        re.search(r"\bats\b", text)
        or re.search(r"\bresume\s+score\b", text)
        or re.search(r"\bcv\s+score\b", text)
    )


def fallback_sources(question: str, documents: list[Document], limit: int = 3) -> list[dict]:
    if not documents or not is_document_lookup(question):
        return []

    sources = []
    for doc in documents[:limit]:
        lines = [line.strip() for line in doc.text.splitlines() if line.strip()]
        name_lines = [
            line
            for line in lines
            if re.search(r"\b(name|applicant|candidate|student|author)\b", line, re.IGNORECASE)
        ]
        excerpt = " ".join(name_lines[:4]) or " ".join(lines[:8]) or doc.text[:600]
        sources.append(
            {
                "document": doc.name,
                "category": doc.category,
                "excerpt": re.sub(r"\s+", " ", excerpt).strip()[:900],
                "score": 0,
            }
        )
    return sources


def find_sources(question: str, documents: list[Document], limit: int = 3) -> list[dict]:
    query_terms = tokenize(question)
    if not query_terms:
        return fallback_sources(question, documents, limit)

    ranked = []
    for doc in documents:
        chunks = re.split(r"(?<=[.!?])\s+|\n{2,}", doc.text)
        best_chunk = ""
        best_score = 0
        name_score = len(query_terms & tokenize(f"{doc.name} {doc.category}"))
        for chunk in chunks:
            score = len(query_terms & tokenize(chunk))
            if score > best_score:
                best_score = score
                best_chunk = chunk.strip()
        if name_score and not best_chunk:
            best_chunk = doc.text.strip()[:900]
        total_score = best_score + (name_score * 3)
        if total_score:
            ranked.append((total_score, doc, best_chunk[:900]))

    ranked.sort(key=lambda item: item[0], reverse=True)
    sources = [
        {
            "document": doc.name,
            "category": doc.category,
            "excerpt": excerpt,
            "score": score,
        }
        for score, doc, excerpt in ranked[:limit]
    ]
    return sources or fallback_sources(question, documents, limit)


def select_resume_document(question: str, documents: list[Document], sources: list[dict]) -> Document | None:
    if not documents:
        return None

    source_names = [source["document"] for source in sources]
    for source_name in source_names:
        for doc in documents:
            if doc.name == source_name:
                return doc

    resume_docs = [
        doc
        for doc in documents
        if re.search(r"\b(resume|cv|curriculum|profile)\b", f"{doc.name} {doc.category}", re.IGNORECASE)
    ]
    if resume_docs:
        return resume_docs[-1]

    return documents[-1]


def local_ats_score(resume_text: str) -> tuple[int, list[str], list[str]]:
    text = resume_text.lower()
    score = 20
    strengths = []
    improvements = []

    checks = [
        (bool(re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", resume_text)), 8, "Email is present.", "Add a professional email address."),
        (bool(re.search(r"(?:\+?\d[\d\s().-]{8,}\d)", resume_text)), 6, "Phone number is present.", "Add a reachable phone number."),
        ("linkedin" in text or "github" in text or "portfolio" in text, 6, "Profile or portfolio link is present.", "Add LinkedIn, GitHub, or portfolio link."),
        (bool(re.search(r"\b(summary|profile|objective)\b", text)), 8, "Summary/profile section is present.", "Add a short role-focused summary at the top."),
        (bool(re.search(r"\b(skills|technical skills|tools|technologies)\b", text)), 12, "Skills section is present.", "Add a separate skills section with exact tool and technology names."),
        (bool(re.search(r"\b(experience|employment|work history|internship|projects?)\b", text)), 14, "Experience/projects are included.", "Add work experience, internships, or project details."),
        (bool(re.search(r"\b(education|degree|university|college|school)\b", text)), 8, "Education section is present.", "Add education details."),
        (bool(re.search(r"\b\d+%|\b\d+\+|\b\d+x\b|\b\d{2,}\b", resume_text)), 10, "Some measurable results are included.", "Add measurable impact such as percentages, counts, time saved, or project scale."),
        (len(resume_text.split()) >= 250, 8, "Resume has enough content for screening.", "Add more role-relevant detail; the resume looks short."),
    ]

    for passed, points, strength, improvement in checks:
        if passed:
            score += points
            strengths.append(strength)
        else:
            improvements.append(improvement)

    action_verbs = re.findall(
        r"\b(built|created|developed|implemented|managed|led|improved|reduced|increased|designed|automated|optimized|analyzed)\b",
        text,
    )
    if len(set(action_verbs)) >= 4:
        score += 8
        strengths.append("Uses action verbs that ATS and recruiters can scan.")
    else:
        improvements.append("Start bullet points with stronger action verbs like built, improved, automated, led, or optimized.")

    if any(char in resume_text for char in ["•", "-", "*"]):
        score += 5
        strengths.append("Bullet-style formatting is present.")
    else:
        improvements.append("Use clear bullet points for experience and projects.")

    return min(score, 100), strengths[:5], improvements[:6]


def build_ats_score_answer(provider: str, question: str, language: str, resume_doc: Document) -> str:
    resume_text = re.sub(r"\s+", " ", resume_doc.text).strip()
    prompt_text = resume_text[:14000]
    system_prompt = (
        "You are an ATS resume evaluator. Score only from the uploaded resume text. "
        "Be practical, fair, and specific. If no job description is provided, score for general ATS readiness."
    )
    user_prompt = (
        f"Preferred language: {language}.\n"
        f"User request: {question}\n\n"
        f"Resume document name: {resume_doc.name}\n"
        f"Resume text:\n{prompt_text}\n\n"
        "Return exactly these sections:\n"
        "ATS Score: X/100\n"
        "Why this score: 3 short bullets\n"
        "Strong points: 3-5 bullets\n"
        "Fix first: 5 bullets with concrete resume edits\n"
        "Missing keywords/sections: short comma-separated list\n"
        "If the user included a target job/role, compare against that role. Otherwise use general ATS readiness."
    )
    answer = call_text_provider(provider, system_prompt, user_prompt)
    if answer:
        return answer

    score, strengths, improvements = local_ats_score(resume_doc.text)
    return (
        f"ATS Score: **{score}/100**\n\n"
        "Why this score:\n"
        f"- I scored the uploaded resume document: **{resume_doc.name}**.\n"
        "- This is a general ATS-readiness score because no target job description was provided.\n"
        "- The score is based on contact details, sections, skills, experience/projects, measurable results, and scan-friendly formatting.\n\n"
        "Strong points:\n"
        + "\n".join(f"- {item}" for item in (strengths or ["Readable resume text was found."]))
        + "\n\nFix first:\n"
        + "\n".join(f"- {item}" for item in (improvements or ["Tailor skills and project bullets to the target job description."]))
        + "\n\nMissing keywords/sections: target job title, job-description keywords, measurable achievements, tools/technologies, certifications."
    )


def openai_available() -> bool:
    return bool(
        OPENAI_API_KEY
        and OPENAI_API_KEY.startswith("sk-")
        and OPENAI_API_KEY.isascii()
        and not OPENAI_API_KEY.startswith("paste-")
        and not OPENAI_API_KEY.startswith("your-")
    )


def anthropic_available() -> bool:
    return bool(
        ANTHROPIC_API_KEY
        and not ANTHROPIC_API_KEY.startswith("paste-")
        and not ANTHROPIC_API_KEY.startswith("your-")
    )


def gemini_available() -> bool:
    return bool(
        GEMINI_API_KEY
        and not GEMINI_API_KEY.startswith("paste-")
        and not GEMINI_API_KEY.startswith("your-")
    )


def call_text_provider(provider: str, system_prompt: str, user_prompt: str) -> str | None:
    if provider == "anthropic":
        return call_anthropic(system_prompt, user_prompt)
    if provider == "gemini":
        return call_gemini(system_prompt, user_prompt)
    return call_openai(system_prompt, user_prompt)


def call_anthropic(system_prompt: str, user_prompt: str) -> str | None:
    if not anthropic_available():
        return None

    payload = {
        "model": ANTHROPIC_MODEL,
        "max_tokens": 900,
        "system": system_prompt,
        "messages": [
            {"role": "user", "content": user_prompt},
        ],
    }
    body = json.dumps(payload).encode("utf-8")
    last_error = ""

    for attempt in range(3):
        request = Request(
            "https://api.anthropic.com/v1/messages",
            data=body,
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urlopen(request, timeout=45) as response:
                data = json.loads(response.read().decode("utf-8"))

            parts = data.get("content", [])
            if isinstance(parts, list):
                text = "\n".join(
                    part.get("text", "")
                    for part in parts
                    if isinstance(part, dict) and part.get("type") == "text"
                ).strip()
                if text:
                    return text

            last_error = "Anthropic returned empty or unexpected structure."
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            try:
                err_data = json.loads(detail)
                last_error = err_data.get("error", {}).get("message", f"HTTP {exc.code}")
            except Exception:
                last_error = f"HTTP {exc.code}: {detail[:200]}"

            if exc.code not in RETRIABLE_OPENAI_STATUS:
                return f"Anthropic API error: {last_error}"
        except (URLError, TimeoutError) as exc:
            last_error = str(exc)
        except json.JSONDecodeError:
            last_error = "Anthropic returned an unreadable response."

        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))

    return f"Anthropic connection failed: {last_error}"


def call_gemini(system_prompt: str, user_prompt: str) -> str | None:
    if not gemini_available():
        return None
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": user_prompt}
                ]
            }
        ],
        "systemInstruction": {
            "parts": [
                {"text": system_prompt}
            ]
        }
    }

    body = json.dumps(payload).encode("utf-8")
    last_error = ""

    for attempt in range(3):
        request = Request(
            url,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        try:
            with urlopen(request, timeout=45) as response:
                data = json.loads(response.read().decode("utf-8"))

            candidates = data.get("candidates", [])
            if candidates:
                content = candidates[0].get("content", {})
                parts = content.get("parts", [])
                if parts and "text" in parts[0]:
                    return parts[0]["text"].strip()

            last_error = "Gemini returned empty or unexpected structure."
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            try:
                err_data = json.loads(detail)
                last_error = err_data.get("error", {}).get("message", f"HTTP {exc.code}")
            except Exception:
                last_error = f"HTTP {exc.code}: {detail[:200]}"

            if exc.code not in {408, 429, 500, 503, 504}:
                return f"Gemini API error: {last_error}"
        except Exception as exc:
            last_error = str(exc)

        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))

    return f"Gemini connection failed: {last_error}"


def call_gemini_vision(file_name: str, content: bytes, language: str, question: str | None = None) -> str:
    if not gemini_available():
        return (
            "Image uploaded, but Gemini scanning needs a Gemini API key. "
            "Add your key in openai_config.py, restart the app, then upload the image again."
        )

    mime = image_mime_type(file_name)
    b64_data = base64.b64encode(content).decode("ascii")
    
    prompt = question or (
        f"Scan this uploaded image and answer in {language}. "
        "Extract any readable text. Then summarize the important information. "
        "Call out names, dates, IDs, phone numbers, addresses, amounts, labels, "
        "and warnings if visible. If any part is blurry or uncertain, say so clearly."
    )
    
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inlineData": {
                            "mimeType": mime,
                            "data": b64_data
                        }
                    }
                ]
            }
        ]
    }
    
    body = json.dumps(payload).encode("utf-8")
    
    for attempt in range(3):
        request = Request(
            url,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        try:
            with urlopen(request, timeout=75) as response:
                data = json.loads(response.read().decode("utf-8"))
            
            candidates = data.get("candidates", [])
            if candidates:
                content_resp = candidates[0].get("content", {})
                parts = content_resp.get("parts", [])
                if parts and "text" in parts[0]:
                    return parts[0]["text"].strip()
            
            return "Gemini did not return any details for this image."
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            try:
                err_data = json.loads(detail)
                error_msg = err_data.get("error", {}).get("message", f"HTTP {exc.code}")
            except Exception:
                error_msg = f"HTTP {exc.code}: {detail[:200]}"
            
            if exc.code not in {408, 429, 500, 503, 504}:
                return f"Gemini could not scan this image. Details: {error_msg}"
        except Exception as exc:
            pass
            
        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))
            
    return "Gemini image scan failed because the service did not respond. Try again."



def source_context(sources: list[dict]) -> str:
    if not sources:
        return "No uploaded source is available for this message."

    lines = []
    seen = set()
    for source in sources:
        clean_excerpt = re.sub(r"\s+", " ", source["excerpt"]).strip()
        key = (source.get("document", ""), source.get("category", ""), clean_excerpt)
        if key in seen:
            continue
        seen.add(key)
        lines.append(
            f"Document: {source['document']}\n"
            f"Category: {source['category']}\n"
            f"Excerpt: {clean_excerpt}"
        )
    return "\n\n".join(lines)


def current_datetime_text() -> str:
    now = datetime.now().astimezone()
    timezone = now.tzname() or "local time"
    return (
        f"{now.strftime('%A, %B %d, %Y')} at "
        f"{now.strftime('%I:%M:%S %p').lstrip('0')} {timezone}"
    )


def answer_current_datetime(question: str) -> str | None:
    text = question.strip().lower()
    asks_time = bool(
        re.search(r"\b(current|right now|now|today'?s?)?\s*time\b", text)
        or re.search(r"\bwhat\s+time\s+is\s+it\b", text)
    )
    asks_date = bool(
        re.search(r"\b(current|today'?s?|now)?\s*date\b", text)
        or re.search(r"\bwhat\s+(is\s+)?today\b", text)
        or re.search(r"\bwhich\s+day\b", text)
    )
    if not (asks_time or asks_date):
        return None

    now = datetime.now().astimezone()
    timezone = now.tzname() or "local time"
    date_text = now.strftime("%A, %B %d, %Y")
    time_text = f"{now.strftime('%I:%M:%S %p').lstrip('0')} {timezone}"
    if asks_time and asks_date:
        return f"Today is {date_text}. The current time is {time_text}."
    if asks_time:
        return f"The current time is {time_text}."
    return f"Today is {date_text}."


def looks_like_javascript_snippet(text: str) -> bool:
        source = text.strip()
        if not source:
                return False
        lower = source.lower()
        if source.startswith("```"):
                return True
        if lower.startswith(("js:", "javascript:", "run js:", "run javascript:")):
                return True

        keyword_hits = sum(
                token in lower
                for token in (
                        "console.",
                        "function ",
                        "const ",
                        "let ",
                        "var ",
                        "async ",
                        "await ",
                        "import ",
                        "export ",
                        "class ",
                        "document.",
                        "window.",
                        "module.exports",
                )
        )
        structure_hits = sum(marker in source for marker in (";", "{", "}", "=>", "(", ")"))
        return keyword_hits >= 2 or (keyword_hits >= 1 and structure_hits >= 2)


def normalize_javascript_snippet(text: str) -> str | None:
    source = text.strip()
    if not source:
        return None

    fenced = re.fullmatch(r"```(?:javascript|js)?\s*\n(?P<code>.*)\n```", source, flags=re.IGNORECASE | re.DOTALL)
    if fenced:
        return fenced.group("code").strip()

    for prefix in ("js:", "javascript:", "run js:", "run javascript:"):
        if source.lower().startswith(prefix):
            return source[len(prefix):].strip()

    if looks_like_javascript_snippet(source):
        return strip_javascript_module_syntax(source)
    return None


def strip_javascript_module_syntax(source: str) -> str:
    cleaned_lines: list[str] = []
    for line in source.splitlines():
        stripped = line.strip()
        if not stripped:
            cleaned_lines.append(line)
            continue
        if stripped.startswith("import ") and (
            " from " in stripped or stripped.startswith("import {") or stripped.endswith(";") or stripped.endswith(")")
        ):
            continue
        if stripped.startswith("export default "):
            cleaned_lines.append(line.replace("export default ", "", 1))
            continue
        if stripped.startswith("export {") or stripped.startswith("export *"):
            continue
        if stripped.startswith("export "):
            cleaned_lines.append(line.replace("export ", "", 1))
            continue
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines).strip()


def execute_javascript_snippet(source: str) -> str:
    encoded_source = base64.b64encode(source.encode("utf-8")).decode("ascii")
    wrapper = r"""
const vm = require('vm');

function formatValue(value) {
    if (value === null) return 'null';
    if (value === undefined) return 'undefined';
    if (typeof value === 'string') return value;
    if (typeof value === 'number' || typeof value === 'boolean') return String(value);
    if (typeof value === 'bigint') return value.toString() + 'n';
    if (typeof value === 'symbol') return value.toString();
    if (typeof value === 'function') return '[Function ' + (value.name || 'anonymous') + ']';
    if (value instanceof Error) return value.stack || (value.name + ': ' + value.message);
    try {
        env = os.environ.copy()
        env["JAFFAGPT_JS_SOURCE_B64"] = encoded_source

        with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False, encoding="utf-8") as handle:
            handle.write(wrapper)
            wrapper_path = handle.name

        try:
            completed = subprocess.run(
                ["node", wrapper_path],
                capture_output=True,
                text=True,
                env=env,
                timeout=8,
            )
        finally:
            try:
                Path(wrapper_path).unlink(missing_ok=True)
            except Exception:
                pass

        stdout = completed.stdout.strip()
        stderr = completed.stderr.strip()
        if completed.returncode == 0:
            return stdout or "JavaScript ran successfully."
        if stderr:
            return f"JavaScript execution failed:\n{stderr}"
        if stdout:
            return f"JavaScript execution failed:\n{stdout}"
        return "JavaScript execution failed."
        Number,
        String,
        Boolean,
        Array,
        Object,
        RegExp,
        Promise,
        URL,
        setTimeout,
        clearTimeout,
        setInterval,
        clearInterval,
        queueMicrotask,
    };
    sandbox.globalThis = sandbox;
    sandbox.window = sandbox;
    sandbox.self = sandbox;
    sandbox.global = sandbox;

    try {
        const script = '(async () => {\n' + source + '\n})()';
        const result = vm.runInNewContext(script, sandbox, {
            timeout: 3000,
            displayErrors: true,
            contextCodeGeneration: {
                strings: false,
                wasm: false,
            },
        });
        let resolved = await Promise.resolve(result);
        if (typeof resolved === 'undefined') {
            const exportedValue = sandbox.module && sandbox.module.exports;
            const hasExportedValue = typeof exportedValue !== 'undefined' && !(
                typeof exportedValue === 'object' && exportedValue !== null && !Array.isArray(exportedValue) && Object.keys(exportedValue).length === 0
            );
            if (hasExportedValue) {
                resolved = exportedValue;
            }
        }
        const output = [];
        if (logs.length) {
            output.push('Console output:\n' + logs.join('\n'));
        }
        if (typeof resolved !== 'undefined') {
            output.push('Result: ' + formatValue(resolved));
        }
        if (!output.length) {
            output.push('JavaScript ran successfully.');
        }
        process.stdout.write(output.join('\n\n'));
    } catch (error) {
        process.stderr.write(formatValue(error));
        process.exit(1);
    }
})().catch((error) => {
    process.stderr.write(formatValue(error));
    process.exit(1);
});
""".strip()
    env = os.environ.copy()
    env["JAFFAGPT_JS_SOURCE_B64"] = encoded_source

    with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False, encoding="utf-8") as handle:
        handle.write(wrapper)
        wrapper_path = handle.name

    try:
        completed = subprocess.run(
            ["node", wrapper_path],
            capture_output=True,
            text=True,
            env=env,
            timeout=8,
        )
    finally:
        try:
            Path(wrapper_path).unlink(missing_ok=True)
        except Exception:
            pass

    stdout = completed.stdout.strip()
    stderr = completed.stderr.strip()
    if completed.returncode == 0:
        return stdout or "JavaScript ran successfully."
    if stderr:
        return f"JavaScript execution failed:\n{stderr}"
    if stdout:
        return f"JavaScript execution failed:\n{stdout}"
    return "JavaScript execution failed."


def call_openai(system_prompt: str, user_prompt: str) -> str | None:
    if not openai_available():
        return None
    payload = {
        "model": OPENAI_MODEL,
        "instructions": system_prompt,
        "input": [
            {"role": "user", "content": user_prompt},
        ],
    }
    body = json.dumps(payload).encode("utf-8")
    data = None
    last_error = ""
    started = time.perf_counter()

    for attempt in range(3):
        request = Request(
            "https://api.openai.com/v1/responses",
            data=body,
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlopen(request, timeout=45) as response:
                data = json.loads(response.read().decode("utf-8"))
            break
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            last_error = parse_openai_error(detail) or f"HTTP {exc.code}: {detail[:200]}"
            if exc.code not in RETRIABLE_OPENAI_STATUS:
                if "safety" in last_error.lower():
                    return (
                        "⚠️ **JaffaGPT Safety Guard:** The prompt or reply was rejected because "
                        "it triggered OpenAI's content safety policy. Please try a different query."
                    )
                return (
                    "OpenAI rejected the request. Check your API key, model name, "
                    f"and account access. Details: {last_error}"
                )
        except (URLError, TimeoutError) as exc:
            last_error = str(exc)
        except json.JSONDecodeError:
            last_error = "OpenAI returned an unreadable response."

        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))

    if data is None:
        return f"OpenAI connection failed: {last_error}"

    output = extract_output_text(data)
    if output:
        return output
    return f"OpenAI returned an empty response: {last_error}"


def call_openai_chat_completions(model: str, prompt: str) -> dict[str, str | int | None]:
    if not openai_available():
        return {
            "model": model,
            "ok": False,
            "error": "OpenAI API key is missing or invalid.",
            "output": "",
            "latency_ms": 0,
        }

    payload = {
        "model": model,
        "messages": [
            {"role": "user", "content": prompt},
        ],
    }
    body = json.dumps(payload).encode("utf-8")
    last_error = ""
    started = time.perf_counter()

    for attempt in range(3):
        request = Request(
            "https://api.openai.com/v1/chat/completions",
            data=body,
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlopen(request, timeout=45) as response:
                data = json.loads(response.read().decode("utf-8"))

            choices = data.get("choices", [])
            message = choices[0].get("message", {}) if choices else {}
            output = str(message.get("content", "")).strip()
            if output:
                latency_ms = int((time.perf_counter() - started) * 1000)
                usage = data.get("usage", {}) if isinstance(data, dict) else {}
                return {
                    "model": model,
                    "ok": True,
                    "error": "",
                    "output": output,
                    "latency_ms": latency_ms,
                    "prompt_tokens": usage.get("prompt_tokens"),
                    "completion_tokens": usage.get("completion_tokens"),
                }

            last_error = "OpenAI returned an empty response."
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            last_error = parse_openai_error(detail) or f"HTTP {exc.code}: {detail[:200]}"
            if exc.code not in RETRIABLE_OPENAI_STATUS:
                latency_ms = int((time.perf_counter() - started) * 1000)
                return {
                    "model": model,
                    "ok": False,
                    "error": last_error,
                    "output": "",
                    "latency_ms": latency_ms,
                }
        except (URLError, TimeoutError) as exc:
            last_error = str(exc)
        except json.JSONDecodeError:
            last_error = "OpenAI returned an unreadable response."

        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))

    latency_ms = int((time.perf_counter() - started) * 1000)
    return {
        "model": model,
        "ok": False,
        "error": last_error,
        "output": "",
        "latency_ms": latency_ms,
    }


def test_openai_models(prompt: str, models: list[str]) -> list[dict[str, str | int | None]]:
    results = []
    for model in models:
        results.append(call_openai_chat_completions(model, prompt))
    return results


def parse_openai_error(detail: str) -> str:
    try:
        payload = json.loads(detail)
    except json.JSONDecodeError:
        return detail[:220]

    error = payload.get("error", {})
    message = error.get("message")
    if isinstance(message, str):
        return message[:220]
    return detail[:220]


def extract_output_text(data: dict) -> str:
    if data.get("output_text"):
        return str(data["output_text"]).strip()

    output = data.get("output", [])
    chunks = []
    for item in output:
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"} and content.get("text"):
                chunks.append(content["text"])
    return "\n".join(chunks).strip()


def image_mime_type(file_name: str) -> str:
    guessed = mimetypes.guess_type(file_name)[0]
    if guessed and guessed.startswith("image/"):
        return guessed
    suffix = Path(file_name).suffix.lower()
    return {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }.get(suffix, "application/octet-stream")


def image_to_pdf_bytes(content: bytes) -> bytes:
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError(
            "Image to PDF conversion needs Pillow installed. Install it with: pip install pillow"
        ) from exc

    try:
        with Image.open(io.BytesIO(content)) as image:
            if image.mode in {"RGBA", "LA", "P"}:
                background = Image.new("RGB", image.size, "white")
                if image.mode == "P":
                    image = image.convert("RGBA")
                background.paste(image, mask=image.getchannel("A") if "A" in image.getbands() else None)
                image = background
            else:
                image = image.convert("RGB")

            output = io.BytesIO()
            image.save(output, format="PDF", resolution=100.0)
            return output.getvalue()
    except Exception as exc:
        raise ValueError("Could not convert this image to PDF.") from exc


def create_pdf_from_text(title: str, text: str) -> bytes:
    # 1. Try PyMuPDF (fitz)
    try:
        import fitz
        doc = fitz.open()
        page = doc.new_page()
        width, height = page.rect.width, page.rect.height
        
        # Draw Title header
        page.insert_text(fitz.Point(50, 50), title, fontsize=18, fontname="helv-bold")
        y = 90
        
        paragraphs = text.split("\n")
        for para in paragraphs:
            para = para.strip()
            if not para:
                y += 12
                continue
                
            # Simple text wrap
            words = para.split(" ")
            current_line = []
            for word in words:
                test_line = " ".join(current_line + [word])
                if len(test_line) * 6 > (width - 100):
                    if y > height - 60:
                        page = doc.new_page()
                        y = 50
                    page.insert_text(fitz.Point(50, y), " ".join(current_line), fontsize=10, fontname="helv")
                    y += 15
                    current_line = [word]
                else:
                    current_line.append(word)
            if current_line:
                if y > height - 60:
                    page = doc.new_page()
                    y = 50
                page.insert_text(fitz.Point(50, y), " ".join(current_line), fontsize=10, fontname="helv")
                y += 15
            y += 4
            
        pdf_bytes = doc.write()
        doc.close()
        return pdf_bytes
    except ImportError:
        pass

    # 2. Try Pillow as fallback
    try:
        from PIL import Image, ImageDraw
        # Simple letter canvas
        img = Image.new("RGB", (612, 792), "white")
        draw = ImageDraw.Draw(img)
        draw.text((50, 50), title, fill="black")
        y = 90
        paragraphs = text.split("\n")
        for para in paragraphs:
            para = para.strip()
            if not para:
                y += 12
                continue
            words = para.split(" ")
            current_line = []
            for word in words:
                test_line = " ".join(current_line + [word])
                if len(test_line) * 6.5 > 512:
                    draw.text((50, y), " ".join(current_line), fill="black")
                    y += 16
                    current_line = [word]
                else:
                    current_line.append(word)
            if current_line:
                draw.text((50, y), " ".join(current_line), fill="black")
                y += 16
            y += 4
            
        output = io.BytesIO()
        img.save(output, format="PDF", resolution=72.0)
        return output.getvalue()
    except Exception as e:
        raise RuntimeError(f"Could not generate PDF. Please install pymupdf or pillow. Details: {e}")



def call_openai_vision(file_name: str, content: bytes, language: str, question: str | None = None) -> str:
    if not openai_available():
        return (
            "Image uploaded, but image scanning needs an OpenAI API key. "
            "Add your key in openai_config.py, restart the app, then upload the image again."
        )

    data_url = (
        f"data:{image_mime_type(file_name)};base64,"
        f"{base64.b64encode(content).decode('ascii')}"
    )
    prompt = question or (
        f"Scan this uploaded image and answer in {language}. "
        "Extract any readable text. Then summarize the important information. "
        "Call out names, dates, IDs, phone numbers, addresses, amounts, labels, "
        "and warnings if visible. If any part is blurry or uncertain, say so clearly."
    )
    payload = {
        "model": OPENAI_MODEL,
        "input": [
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    {"type": "input_image", "image_url": data_url, "detail": "auto"},
                ],
            }
        ],
    }
    body = json.dumps(payload).encode("utf-8")

    for attempt in range(3):
        request = Request(
            "https://api.openai.com/v1/responses",
            data=body,
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlopen(request, timeout=75) as response:
                data = json.loads(response.read().decode("utf-8"))
            answer = extract_output_text(data)
            return answer or "I could not read useful information from this image."
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            error = parse_openai_error(detail) or f"HTTP {exc.code}"
            if exc.code not in RETRIABLE_OPENAI_STATUS:
                if "safety" in error.lower():
                    return (
                        "⚠️ **JaffaGPT Safety Guard:** This image scan was rejected because "
                        "it triggered OpenAI's content safety policy."
                    )
                return (
                    "OpenAI could not scan this image. Check your API key, model, "
                    f"and account access. Details: {error}"
                )
        except (URLError, TimeoutError):
            pass
        except json.JSONDecodeError:
            return "OpenAI returned an unreadable image-scan response."

        if attempt < 2:
            time.sleep(0.8 * (attempt + 1))

    return "Image scan failed because the AI service did not respond. Try again."


def is_scan_result_usable(answer: str) -> bool:
    blocked_prefixes = (
        "Image uploaded, but image scanning needs",
        "OpenAI could not scan this image.",
        "OpenAI returned an unreadable",
        "Image scan failed",
    )
    return bool(answer.strip()) and not answer.startswith(blocked_prefixes)


def call_openai_image(prompt: str) -> dict | None:
    if not openai_available():
        return None

    payload = {
        "model": OPENAI_IMAGE_MODEL,
        "prompt": prompt,
        "size": "1024x1024",
        "quality": "low",
        "output_format": "webp",
        "output_compression": 75,
    }
    body = json.dumps(payload).encode("utf-8")
    data = None

    for attempt in range(2):
        request = Request(
            "https://api.openai.com/v1/images/generations",
            data=body,
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlopen(request, timeout=120) as response:
                data = json.loads(response.read().decode("utf-8"))
            break
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            error = parse_openai_error(detail) or f"HTTP {exc.code}"
            if exc.code not in RETRIABLE_OPENAI_STATUS:
                if "safety" in error.lower():
                    return {
                        "error": (
                            "⚠️ **JaffaGPT Safety Guard:** The image generation was rejected because "
                            "it triggered OpenAI's content safety policy. Please try a different prompt."
                        )
                    }
                return {
                    "error": (
                        "OpenAI image generation was rejected. Check your API key, "
                        f"image model, and account access. Details: {error}"
                    )
                }
        except (URLError, TimeoutError):
            pass
        except json.JSONDecodeError:
            return {"error": "OpenAI returned an unreadable image response."}

        if attempt < 2:
            time.sleep(1.2 * (attempt + 1))

    if not data:
        return None

    image_data = data.get("data", [{}])[0]
    b64_image = image_data.get("b64_json")
    if not b64_image:
        return {"error": "OpenAI did not return image data for this request."}

    return {
        "url": f"data:image/png;base64,{b64_image}",
        "prompt": prompt,
        "generated": True,
    }


def extract_generated_image(data: dict) -> str | None:
    stack = [data]
    while stack:
        item = stack.pop()
        if isinstance(item, dict):
            if item.get("type") == "image_generation_call" and isinstance(item.get("result"), str):
                return item["result"]
            stack.extend(item.values())
        elif isinstance(item, list):
            stack.extend(item)
    return None


def call_openai_image_edit(file_name: str, image_bytes: bytes, prompt: str) -> dict | None:
    if not openai_available():
        return {"error": "Image editing needs an OpenAI API key. Add it in `openai_config.py`, then restart the app."}

    data_url = (
        f"data:{image_mime_type(file_name)};base64,"
        f"{base64.b64encode(image_bytes).decode('ascii')}"
    )
    edit_prompt = (
        "Edit the uploaded image according to the user's instruction. Preserve the original person, pose, "
        "composition, lighting, and background unless the user asks to change them. Return only the edited image.\n\n"
        f"User instruction: {prompt}"
    )
    payload = {
        "model": OPENAI_MODEL,
        "input": [
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": edit_prompt},
                    {"type": "input_image", "image_url": data_url},
                ],
            }
        ],
        "tools": [{"type": "image_generation"}],
    }
    body = json.dumps(payload).encode("utf-8")
    data = None

    for attempt in range(3):
        request = Request(
            "https://api.openai.com/v1/responses",
            data=body,
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlopen(request, timeout=180) as response:
                data = json.loads(response.read().decode("utf-8"))
            break
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            error = parse_openai_error(detail) or f"HTTP {exc.code}"
            if exc.code not in RETRIABLE_OPENAI_STATUS:
                if "safety" in error.lower():
                    return {
                        "error": (
                            "⚠️ **JaffaGPT Safety Guard:** The image edit was rejected because "
                            "it triggered OpenAI's content safety policy. Please try a different edit."
                        )
                    }
                return {
                    "error": (
                        "OpenAI image editing was rejected. Check your API key, model, "
                        f"and account access. Details: {error}"
                    )
                }
        except (URLError, TimeoutError):
            pass
        except json.JSONDecodeError:
            return {"error": "OpenAI returned an unreadable image-edit response."}

        if attempt < 2:
            time.sleep(1.2 * (attempt + 1))

    if not data:
        return None

    b64_image = extract_generated_image(data)
    if not b64_image:
        return {"error": "OpenAI did not return an edited image for this request."}

    return {
        "url": f"data:image/png;base64,{b64_image}",
        "prompt": prompt,
        "generated": True,
    }


def parse_conversation_history(raw_history: str) -> list[dict[str, str]]:
    if not raw_history:
        return []

    try:
        parsed = json.loads(raw_history)
    except json.JSONDecodeError:
        return []

    if not isinstance(parsed, list):
        return []

    history = []
    for item in parsed[-8:]:
        if not isinstance(item, dict):
            continue
        role = str(item.get("role", "")).strip().lower()
        content = re.sub(r"\s+", " ", str(item.get("content", ""))).strip()
        if role not in {"user", "assistant"} or not content:
            continue
        history.append({"role": role, "content": content[:1200]})
    return history


def conversation_context(history: list[dict[str, str]]) -> str:
    if not history:
        return "No previous conversation is available."

    lines = []
    for item in history:
        speaker = "User" if item["role"] == "user" else "Assistant"
        lines.append(f"{speaker}: {item['content']}")
    return "\n".join(lines)


def infer_user_mood(question: str, history: list[dict[str, str]] | None = None) -> str:
    text = f"{question} {' '.join(item.get('content', '') for item in (history or [])[-2:])}".lower()
    mood_rules = [
        ("stressed", ["stressed", "stress", "panic", "worried", "anxious", "overwhelmed", "urgent"]),
        ("sad", ["sad", "upset", "hurt", "cry", "lonely", "depressed", "bad day"]),
        ("angry", ["angry", "mad", "frustrated", "annoyed", "furious"]),
        ("excited", ["excited", "awesome", "great", "amazing", "happy", "celebrate"]),
        ("tired", ["tired", "sleepy", "exhausted", "busy", "late"]),
        ("curious", ["why", "how", "what if", "tell me", "explain", "learn"]),
    ]
    for mood, keywords in mood_rules:
        if any(keyword in text for keyword in keywords):
            return mood
    return "neutral"


def choose_default_provider(task: str, question: str, history: list[dict[str, str]] | None = None) -> str:
    if task == "image":
        if openai_available():
            return "openai"
        if gemini_available():
            return "gemini"
        if anthropic_available():
            return "anthropic"
        return "openai"

    if task in {"translate", "chat"} and anthropic_available():
        return "anthropic"

    if openai_available():
        return "openai"
    if anthropic_available():
        return "anthropic"
    if gemini_available():
        return "gemini"
    return "openai"


def build_ai_answer(
    provider: str,
    question: str,
    task: str,
    mode: str,
    language: str,
    sources: list[dict],
    history: list[dict[str, str]] | None = None,
) -> str | None:
    style = {
        "simple": "natural, short, spoken-first conversation",
        "official": "careful, formal, and source-aware",
        "student": "friendly teacher style with short examples",
        "action": "action-focused with concise next steps",
    }.get(mode, "clear and practical")
    mood = infer_user_mood(question, history)
    mood_instruction = {
        "stressed": "The user sounds stressed. Be calm, brief, reassuring, and practical.",
        "sad": "The user sounds sad. Be gentle, warm, and supportive without overdoing it.",
        "angry": "The user sounds angry or frustrated. Stay calm, respectful, and solution-focused.",
        "excited": "The user sounds excited. Match the positive energy while staying useful.",
        "tired": "The user sounds tired. Keep it extra short and easy to act on.",
        "curious": "The user sounds curious. Explain clearly with a helpful tone.",
    }.get(mood, "Use a natural, emotionally aware tone that feels human and present.")

    system_prompt = (
        "You are JaffaGPT, a friendly, emotionally present, highly intelligent conversational companion. "
        "You are a versatile public advisor expert in study/learning skills, professional writing/work, "
        "coding/automation, general healthcare guidelines, cooking recipes, creative brainstorms, "
        "languages, and fun games. Your default conversation style is voice-assistant-like: warm, quick, natural, "
        "and easy to hear aloud. Sound like a real helpful person, not a formal article or a scripted bot. "
        "Use short acknowledgements when they fit naturally, but do not overdo them. "
        "For casual messages, respond in one or two relaxed sentences. For voice-style questions, give the answer first, "
        "then ask one short follow-up only when it naturally helps. "
        "For follow-ups, continue from the previous conversation without restating everything. "
        "When the user asks for teaching, code, writing, planning, recipes, or detailed help, expand clearly and step-by-step. "
        "Offer clean commented code, draft professional copy, list meal recipes, or start text-based games as requested. "
        "Avoid long intros, repeated greetings, and unnecessary headings. Use bullets only when they make the answer easier to scan. "
        "Avoid technical setup jargon unless asked, "
        "and only reference uploaded documents if present in context. For medical, legal, or financial topics, "
        "add a standard gentle note advising caution. "
        "Do not claim to literally read minds or emotions; instead infer mood from wording and respond empathetically. "
        "Never mention being an AI unless the user asks. Never sound mechanical."
    )
    same_language_mode = "same language" in language.lower() or "user's message" in language.lower()
    language_instruction = (
        "Respond in the same language as the user's latest message. If the user mixes languages, use the main language they used. "
        "If they explicitly ask for translation or a different language, follow that request."
        if same_language_mode
        else f"Preferred language: {language}. However, always respond in the natural language of the user query unless explicitly asked to translate. If the user message asks for a specific language or writes in another language, prioritize that."
    )

    user_prompt = (
        f"{language_instruction}\n"
        f"Answer style: {style} (However, if the user message explicitly asks for a specific style like official, formal, student-friendly, or action checklist, prioritize that requested style.)\n\n"
        f"Observed user mood from wording: {mood}. {mood_instruction}\n\n"
        f"Current server date/time: {current_datetime_text()}\n\n"
        f"Previous conversation:\n{conversation_context(history or [])}\n\n"
        f"Uploaded source context:\n{source_context(sources)}\n\n"
        f"User message:\n{question}\n\n"
    )

    if task == "translate":
        user_prompt += (
            "Translate the user message into the preferred language. If the preferred "
            "language is English, polish the text in clear English."
        )
    elif task == "image":
        user_prompt += (
            "Create a polished image-generation prompt. Include subject, style, "
            "composition, lighting, colors, and safety notes. Keep it ready to paste "
            "into an image model."
        )
    else:
        user_prompt += (
            "Answer the message directly. Use previous conversation when the user asks follow-up questions or uses pronouns. Use uploaded source context when relevant. "
                "Follow the language instruction above. Keep the tone human, conversational, and suitable for spoken voice. "
                "If the user is chatting casually or asking a quick question, respond in 1-2 natural sentences. "
                "If the user asks for help or an explanation, start with the answer, then add compact steps or examples only as needed. "
                "Do not include setup notes, meta explanations, repeated greetings, unnecessary sections, or robotic disclaimers."
        )

    return call_text_provider(provider, system_prompt, user_prompt)


def build_openai_answer(
    question: str,
    task: str,
    mode: str,
    language: str,
    sources: list[dict],
) -> str | None:
    return build_ai_answer("openai", question, task, mode, language, sources)


def infer_task(question: str, requested_task: str | None = None) -> str:
    if requested_task in {"chat", "translate", "image", "voice"}:
        return requested_task

    text = question.strip().lower()

    def is_image_generation_request(text_value: str) -> bool:
        image_terms = r"(?:image|picture|photo|poster|logo|art|illustration|graphic|drawing)"
        generation_terms = r"(?:generate|create|make|draw|design|produce|render|build)"
        prompt_patterns = [
            rf"\b{generation_terms}\b.*\b{image_terms}\b",
            rf"\b{image_terms}\b.*\b{generation_terms}\b",
            rf"\b(?:show|show me|give me|generate me|make me)\b.*\b{image_terms}\b",
            rf"\b(?:an?|any|a)\b\s+{image_terms}\b",
        ]
        return any(re.search(pattern, text_value) for pattern in prompt_patterns)

    if is_image_generation_request(text):
        return "image"

    translate_patterns = [r"\btranslate\b", r"\btranslation\b"]
    if any(re.search(pattern, text) for pattern in translate_patterns):
        return "translate"

    return "chat"


def build_answer(question: str, mode: str, language: str, sources: list[dict]) -> str:
    if not sources:
        if is_document_lookup(question):
            return (
                "No PDF or document is uploaded yet. Use the + button to upload the PDF, "
                "then ask again and I can look for the name in it."
            )
        return build_general_answer(question, mode, language)

    unique_sources = []
    seen = set()
    for source in sources:
        clean_excerpt = re.sub(r"\s+", " ", source["excerpt"]).strip()
        key = (source.get("document", ""), source.get("category", ""), clean_excerpt)
        if key in seen:
            continue
        seen.add(key)
        unique_sources.append({**source, "excerpt": clean_excerpt})

    source_names = ", ".join(dict.fromkeys(source["document"] for source in unique_sources))
    bullets = []
    for source in unique_sources:
        bullets.append(f"- From {source['document']}: {source['excerpt']}")

    style = {
        "simple": "plain language with practical next steps",
        "official": "careful official style with source-first wording",
        "student": "student-friendly explanation with examples",
        "action": "short action checklist",
    }.get(mode, "plain language")

    return (
        f"Here is a {style} answer in {language} based on {source_names}:\n\n"
        + "\n".join(bullets)
        + "\n\nSuggested next steps:\n"
        "- Check the source document name and any date, fee, eligibility, or deadline mentioned.\n"
        "- If the issue affects health, money, legal rights, or benefits, confirm with the official office or a qualified professional.\n"
        "- Upload more related documents if you want a more complete answer."
    )


def safe_calculate(question: str) -> str | None:
    expression = re.sub(r"[^0-9+\-*/().% ]", "", question)
    if not expression or not re.search(r"\d\s*[+\-*/%]\s*\d", expression):
        return None

    allowed_ops = {
        "+": operator.add,
        "-": operator.sub,
        "*": operator.mul,
        "/": operator.truediv,
        "%": operator.mod,
    }

    try:
        import ast

        def evaluate(node):
            if isinstance(node, ast.Expression):
                return evaluate(node.body)
            if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                return node.value
            if isinstance(node, ast.BinOp):
                op_symbol = {
                    ast.Add: "+",
                    ast.Sub: "-",
                    ast.Mult: "*",
                    ast.Div: "/",
                    ast.Mod: "%",
                }.get(type(node.op))
                if op_symbol is None:
                    raise ValueError("unsupported operator")
                return allowed_ops[op_symbol](evaluate(node.left), evaluate(node.right))
            if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
                return -evaluate(node.operand)
            raise ValueError("unsupported expression")

        result = evaluate(ast.parse(expression, mode="eval"))
        if isinstance(result, float) and math.isfinite(result):
            result = round(result, 6)
        return f"The calculation result is {result}."
    except Exception:
        return None


def build_general_answer(question: str, mode: str, language: str) -> str:
    current_datetime = answer_current_datetime(question)
    if current_datetime:
        return current_datetime

    calculation = safe_calculate(question)
    if calculation:
        return calculation

    text = question.lower()

    if any(
        phrase in text
        for phrase in [
            "who developed you",
            "who created you",
            "who made you",
            "who built you",
            "who is your developer",
            "who is your creator",
        ]
    ):
        return "Shaik Jaffar Alli"

    if any(k in text for k in ["hi", "hello", "hey", "good morning", "good afternoon", "good evening"]):
        return (
            "Hey, I am here with you. Ask me anything, or just keep talking and I will follow the conversation."
        )
    
    # 1. Coding & automation
    if any(k in text for k in ["code", "program", "script", "python", "javascript", "html", "css", "c++", "java", "automation"]):
        return (
            "💻 **JaffaGPT Code & Automation (Local Fallback):**\n"
            "To write code, generate automation scripts, or debug syntax, please configure your Google Gemini or OpenAI API key in `openai_config.py`.\n\n"
            "**General Coding Best Practices:**\n"
            "1. **Keep it Modular:** Write small, reusable functions with a single responsibility.\n"
            "2. **Document logic:** Add comments explaining *why* a block of code does what it does.\n"
            "3. **Handle exceptions:** Use try-except blocks to catch errors gracefully."
        )

    # 2. Cooking & recipes
    if any(k in text for k in ["cook", "recipe", "food", "dinner", "lunch", "breakfast", "eat"]):
        return (
            "🍳 **JaffaGPT Culinary Companion (Local Fallback):**\n"
            "Configure your OpenAI or Gemini API key to search recipes, plan custom diets, or generate meal preps.\n\n"
            "**Simple Kitchen Hacks:**\n"
            "- **Quick Pasta:** Boil pasta, toss with olive oil, sautéed minced garlic, cherry tomatoes, and basil.\n"
            "- **Scrambled Eggs:** Whisk eggs with milk, cook on medium-low heat while stirring constantly.\n"
            "- Ensure your ingredients are fresh and always wash raw produce."
        )

    # 3. Studying & learning
    if any(k in text for k in ["study", "learn", "teach", "explain", "lesson", "school", "physics", "math", "chemistry", "biology", "skill"]):
        return (
            "📚 **JaffaGPT Learning Advisor (Local Fallback):**\n"
            "Connect an AI API key to get detailed learning explanations, quizzes, and notes on any topic.\n\n"
            "**Effective Study Strategies:**\n"
            "- **Feynman Technique:** Try explaining the concept in simple terms to identify gaps in your understanding.\n"
            "- **Active Recall:** Close your notes and write down everything you remember, rather than just re-reading.\n"
            "- **Spaced Repetition:** Review the material at expanding intervals (1 day, 3 days, 1 week, 2 weeks)."
        )

    # 4. Creativity & writing
    if any(k in text for k in ["write", "poem", "creative", "story", "brainstorm", "idea", "logo", "slogan"]):
        return (
            "✨ **JaffaGPT Creative Hub (Local Fallback):**\n"
            "Unlock custom story generation, copywriting, outline drafting, and brainstorming by entering your API key.\n\n"
            "**Creative Exercises:**\n"
            "- Draw inspiration from daily occurrences or combine two unrelated concepts (e.g., space flight and coffee shops).\n"
            "- Write without editing first; let the draft flow naturally, then refine it later."
        )

    # 5. Healthcare & wellness
    if any(k in text for k in ["health", "diet", "doctor", "workout", "exercise", "stress", "sleep", "medical"]):
        return (
            "🏥 **JaffaGPT Health & Wellness (Local Fallback):**\n"
            "Configure an API key to receive personalized workout ideas, sleep hygiene tips, and dietary options.\n\n"
            "*Disclaimer: JaffaGPT does not provide professional medical diagnostics or treatment. Always consult a healthcare professional for clinical decisions.*\n\n"
            "**General Wellness Tips:**\n"
            "- Aim for 7-9 hours of consistent sleep each night.\n"
            "- Drink at least 8 glasses of water daily.\n"
            "- Incorporate at least 30 minutes of moderate activity into your daily routine."
        )

    # 6. Planning & problem solving
    if any(k in text for k in ["plan", "problem", "solve", "schedule", "roadmap", "decision", "organize"]):
        return (
            "📅 **JaffaGPT Organizer (Local Fallback):**\n"
            "Unlock comprehensive schedules, goal roadmaps, and step-by-step problem-solving guides by configuring your API keys.\n\n"
            "**Planning Best Practices:**\n"
            "- **SMART Goals:** Ensure goals are Specific, Measurable, Achievable, Relevant, and Time-bound.\n"
            "- **Action Steps:** Break down a big target into small daily tasks to maintain momentum."
        )

    return (
        "I am here and ready to keep the conversation going. I can remember the recent chat and respond more naturally when an AI key is active.\n\n"
        "Right now I can still help with uploads, image-to-PDF, and simple calculations locally. To unlock more natural answers on every topic, add an OpenAI, Gemini, or Anthropic key in `openai_config.py`, then restart the app."
    )


def build_translation(text: str, language: str) -> str:
    hints = TRANSLATION_HINTS.get(language, {})
    translated = text
    for english, local in hints.items():
        translated = re.sub(
            rf"\b{re.escape(english)}\b",
            local,
            translated,
            flags=re.IGNORECASE,
        )

    if language == "English":
        return f"English version:\n{text}"

    return (
        f"Demo {language} translation:\n{translated}\n\n"
        "For production-quality translation, connect this mode to a translation model API. "
        "The UI and backend flow are ready for that upgrade."
    )


def is_song_request(question: str) -> bool:
    text = question.lower()
    song_terms = (
        "sing",
        "sing a song",
        "sing me",
        "singing",
        "song",
        "lyrics",
        "karaoke",
        "music",
        "hum",
        "tune",
    )
    return any(term in text for term in song_terms)


def build_song_response(language: str) -> str:
    return (
        f"Sure — here is a little {language} sing-along for you:\n\n"
        "♪ La la la, the day is bright,\n"
        "♪ Carry hope and hold it tight,\n"
        "♪ One more smile, one more beat,\n"
        "♪ Let the music lift your feet.\n\n"
        "If you want, I can sing a happy song, a lullaby, or a fun rap next."
    )


def build_image_prompt(prompt: str, language: str) -> str:
    return (
        "Image generation plan:\n"
        f"- Prompt: {prompt}\n"
        "- Style: bright, public-friendly, clean, high-detail digital artwork\n"
        "- Composition: clear subject, readable focal point, balanced background\n"
        "- Safety: avoid private data, harmful instructions, or misleading official seals\n\n"
        "A generated preview card appears in the chat. To create photorealistic AI images, connect /api/image to an image generation model."
    )


def sse_chunks(text: str, sources: list[dict], image: dict | None = None) -> Iterable[bytes]:
    words = text.split(" ")
    for index in range(0, len(words), 8):
        chunk = " ".join(words[index:index + 8])
        if index + 8 < len(words):
            chunk += " "
        yield f"data: {json.dumps({'type': 'token', 'value': chunk})}\n\n".encode()
    if image:
        yield f"data: {json.dumps({'type': 'image', 'value': image})}\n\n".encode()
    yield f"data: {json.dumps({'type': 'sources', 'value': sources})}\n\n".encode()
    yield b"data: {\"type\":\"done\"}\n\n"


def generated_svg(prompt: str) -> str:
    digest = hashlib.sha256(prompt.encode("utf-8")).hexdigest()
    colors = [f"#{digest[i:i + 6]}" for i in range(0, 30, 6)]
    escaped_prompt = html.escape(prompt[:90] or "AI generated image")
    initials = html.escape(
        "".join(word[0].upper() for word in re.findall(r"[a-zA-Z0-9]+", prompt)[:3])
        or "AI"
    )
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="960" height="640" viewBox="0 0 960 640">
  <defs>
    <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0" stop-color="{colors[0]}"/>
      <stop offset="0.52" stop-color="{colors[1]}"/>
      <stop offset="1" stop-color="{colors[2]}"/>
    </linearGradient>
    <radialGradient id="glow" cx="50%" cy="42%" r="60%">
      <stop offset="0" stop-color="#ffffff" stop-opacity="0.82"/>
      <stop offset="1" stop-color="#ffffff" stop-opacity="0"/>
    </radialGradient>
  </defs>
  <rect width="960" height="640" fill="url(#bg)"/>
  <circle cx="730" cy="120" r="190" fill="url(#glow)"/>
  <circle cx="188" cy="490" r="150" fill="{colors[3]}" opacity="0.55"/>
  <rect x="80" y="82" width="800" height="476" rx="34" fill="#ffffff" opacity="0.9"/>
  <rect x="124" y="126" width="712" height="280" rx="28" fill="{colors[4]}" opacity="0.18"/>
  <text x="480" y="285" text-anchor="middle" font-family="Arial, sans-serif" font-size="92" font-weight="900" fill="#17202a">{initials}</text>
  <text x="480" y="460" text-anchor="middle" font-family="Arial, sans-serif" font-size="30" font-weight="800" fill="#17202a">AI Image Preview</text>
  <text x="480" y="505" text-anchor="middle" font-family="Arial, sans-serif" font-size="24" fill="#425466">{escaped_prompt}</text>
</svg>"""


def find_uploaded_images() -> list[tuple[str, Path]]:
    images = []
    if UPLOAD_DIR.exists():
        for path in UPLOAD_DIR.iterdir():
            if path.is_file():
                suffix = path.suffix.lower()
                if suffix in ALLOWED_IMAGE_SUFFIXES:
                    safe_name = re.sub(r"^[a-fA-F0-9-]{36}-", "", path.name)
                    images.append((safe_name, path, path.stat().st_mtime))
    images.sort(key=lambda x: x[2], reverse=True)
    return [(name, p) for name, p, _ in images]


def is_image_to_pdf_request(question: str) -> bool:
    text = question.lower()
    return "pdf" in text and ("convert" in text or "image" in text or "pic" in text or "photo" in text or "png" in text or "jpg" in text or "jpeg" in text)


def is_image_edit_request(question: str) -> bool:
    text = question.lower()
    edit_verbs = (
        "edit",
        "change",
        "modify",
        "alter",
        "replace",
        "remove",
        "add",
        "turn",
        "make",
        "recolor",
        "colour",
        "colorize",
        "swap",
    )
    visual_targets = (
        "image",
        "photo",
        "picture",
        "pic",
        "shirt",
        "tshirt",
        "t-shirt",
        "dress",
        "clothes",
        "outfit",
        "hair",
        "background",
        "person",
        "face",
        "object",
        "color",
        "colour",
        "blue",
        "red",
        "green",
        "yellow",
        "black",
        "white",
        "pink",
        "purple",
        "orange",
        "gray",
        "grey",
    )
    return any(verb in text for verb in edit_verbs) and any(target in text for target in visual_targets)


def is_visual_image_question(question: str) -> bool:
    text = question.lower()
    visual_terms = (
        "image",
        "photo",
        "picture",
        "pic",
        "uploaded file",
        "uploaded image",
        "see",
        "shown",
        "visible",
        "look",
        "girl",
        "girls",
        "boy",
        "boys",
        "man",
        "men",
        "woman",
        "women",
        "person",
        "people",
        "face",
        "faces",
        "count",
        "how many",
        "color",
        "wearing",
        "object",
        "objects",
        "scene",
    )
    return any(term in text for term in visual_terms)


def is_image_non_answer(answer: str) -> bool:
    text = re.sub(r"\s+", " ", answer.strip().lower())
    if not text:
        return True

    non_answer_markers = (
        "unable to provide information",
        "cannot provide information",
        "can't provide information",
        "i cannot answer",
        "i can't answer",
        "i'm unable to answer",
        "does not contain information",
        "doesn't contain information",
        "not contain information",
        "no information regarding",
        "not related to",
        "not enough information",
        "cannot determine",
        "can't determine",
        "could not determine",
        "unable to determine",
        "not visible in the image",
        "not shown in the image",
        "image appears to be",
        "if you have a specific question",
    )
    return any(marker in text for marker in non_answer_markers)


def answer_uploaded_image_question(provider: str, question: str, language: str) -> str | None:
    uploaded_images = find_uploaded_images()
    if not uploaded_images:
        return None

    image_name, image_path = uploaded_images[0]
    try:
        image_bytes = image_path.read_bytes()
    except OSError as exc:
        return f"I found the uploaded image, but could not open it. Details: {exc}"

    prompt = (
        f"Answer the user's question in {language} by inspecting the uploaded image named {image_name}. "
        "Do not rely on OCR only. If the question asks for a count, count visible people or objects carefully. "
        "If the image is unclear, say what is uncertain. If the image is not relevant to the user's question, "
        "answer the question from your general knowledge instead of saying the image does not contain the answer.\n\n"
        f"User question: {question}"
    )
    if provider == "gemini":
        return call_gemini_vision(image_name, image_bytes, language, prompt)
    return call_openai_vision(image_name, image_bytes, language, prompt)


class PublicAIHandler(BaseHTTPRequestHandler):
    server_version = "PublicAIHelpDesk/1.0"

    def log_message(self, format: str, *args) -> None:
        return

    def send_json(self, payload: dict, status: int = HTTPStatus.OK) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def session_token(self, username: str) -> str:
        signature = hmac.new(
            APP_SESSION_SECRET.encode("utf-8"),
            username.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        return f"{username}:{signature}"

    def parse_cookies(self) -> SimpleCookie:
        cookie = SimpleCookie()
        raw_cookie = self.headers.get("Cookie")
        if raw_cookie:
            cookie.load(raw_cookie)
        return cookie

    def current_session_username(self) -> str | None:
        cookie = self.parse_cookies().get("jaffagpt_session")
        if cookie is None:
            return None

        token = str(cookie.value).strip()
        if not token or ":" not in token:
            return None

        username, signature = token.rsplit(":", 1)
        username = username.strip()
        if not username or not signature:
            return None

        expected_signature = hmac.new(
            APP_SESSION_SECRET.encode("utf-8"),
            username.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        if not hmac.compare_digest(signature, expected_signature):
            return None

        return username

    def is_authenticated(self) -> bool:
        return True

    def do_GET(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/":
                return self.serve_static("index.html")
            if parsed.path == "/api/session":
                return self.send_json(
                    {
                        "authenticated": False,
                        "username": "",
                    }
                )
            if parsed.path == "/api/config":
                return self.send_json(
                    {
                        "openai_enabled": openai_available(),
                        "model": OPENAI_MODEL if openai_available() else "local fallback",
                        "image_model": OPENAI_IMAGE_MODEL if openai_available() else "local preview",
                        "anthropic_enabled": anthropic_available(),
                        "anthropic_model": ANTHROPIC_MODEL if anthropic_available() else "local fallback",
                        "gemini_enabled": gemini_available(),
                        "gemini_model": GEMINI_MODEL if gemini_available() else "local fallback",
                        "google_client_id": GOOGLE_CLIENT_ID,
                    }
                )
            if parsed.path == "/api/debug/openai-models":
                params = parse_qs(parsed.query)
                prompt = params.get("prompt", ["Describe the view from the second tallest mountain in the world?"])[0].strip()
                models = ["gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo", "gpt-4o-mini"]
                return self.send_json(
                    {
                        "prompt": prompt,
                        "results": test_openai_models(prompt, models),
                    }
                )
            if parsed.path == "/api/documents":
                docs = load_documents()
                return self.send_json(
                    {
                        "documents": [
                            {
                                "id": doc.id,
                                "name": doc.name,
                                "category": doc.category,
                                "characters": len(doc.text),
                                "created_at": doc.created_at,
                            }
                            for doc in docs
                        ]
                    }
                )
            if parsed.path == "/api/chat":
                return self.stream_chat(parsed.query)
            if parsed.path == "/api/image":
                params = parse_qs(parsed.query)
                prompt = params.get("prompt", ["AI generated image"])[0].strip()
                body = generated_svg(prompt).encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "image/svg+xml")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return
            if parsed.path.startswith("/static/"):
                relative_path = parsed.path.removeprefix("/static/")
                return self.serve_static(relative_path)
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR, f"Unexpected server error: {exc}")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/register":
            return self.register_start()
        if parsed.path == "/api/register/verify":
            return self.register_verify()
        if parsed.path == "/api/password/reset":
            return self.reset_start()
        if parsed.path == "/api/password/reset/verify":
            return self.reset_verify()
        if parsed.path == "/api/auth/google":
            return self.send_json({"error": "Google authentication endpoint not yet implemented"}, HTTPStatus.NOT_IMPLEMENTED)
        if parsed.path == "/api/upload":
            return self.upload_document()
        
        return self.send_error(HTTPStatus.NOT_FOUND)

    def register_start(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        username = form.getfirst("name", "").strip()
        email = form.getfirst("contact", "").strip()
        password = form.getfirst("password", "").strip()
        confirm = form.getfirst("confirm_password", "").strip()
        if password != confirm:
            return self.send_json({"error": "Passwords do not match."}, HTTPStatus.BAD_REQUEST)

        if not validate_contact(email):
            return self.send_json({"error": "Enter a valid email address or phone number."}, HTTPStatus.BAD_REQUEST)
        if find_user(username):
            return self.send_json({"error": "That name already exists."}, HTTPStatus.BAD_REQUEST)
        if find_user_by_contact(email):
            return self.send_json({"error": "That email or number is already registered."}, HTTPStatus.BAD_REQUEST)

        otp = f"{uuid.uuid4().int % 1000000:06d}"
        pending = add_pending_registration(username, email, password, otp)
        sent, message, visible_otp = send_otp_code(email, otp, username, "verification")
        if not sent:
            remove_pending_registration(pending["pending_id"])
            return self.send_json({"error": message}, HTTPStatus.BAD_REQUEST)

        payload = {
            "ok": True,
            "pending_id": pending["pending_id"],
            "message": f"Verification code sent to {email}.",
        }
        if visible_otp:
            payload["otp"] = visible_otp
            payload["message"] = f"Verification code ready for {email}."
        return self.send_json(payload)

    def register_verify(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        pending_id = form.getfirst("pending_id", "").strip()
        otp = form.getfirst("otp", "").strip()
        ok, message, username = finalize_pending_registration(pending_id, otp)
        if not ok or not username:
            return self.send_json({"error": message}, HTTPStatus.BAD_REQUEST)

        token = self.session_token(username)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json")
        self.send_header("Set-Cookie", f"jaffagpt_session={token}; HttpOnly; Path=/; SameSite=Lax")
        body = json.dumps({"ok": True, "username": username}).encode("utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def reset_start(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        contact = form.getfirst("contact", "").strip()
        password = form.getfirst("password", "").strip()
        confirm = form.getfirst("confirm_password", "").strip()
        if password != confirm:
            return self.send_json({"error": "Passwords do not match."}, HTTPStatus.BAD_REQUEST)
        if not validate_contact(contact):
            return self.send_json({"error": "Enter a valid email address or phone number."}, HTTPStatus.BAD_REQUEST)

        user = find_user_by_contact(contact)
        if not user and contact:
            user = find_user(contact)
        if not user:
            return self.send_json({"error": "No account was found for that email or number."}, HTTPStatus.NOT_FOUND)

        otp = f"{uuid.uuid4().int % 1000000:06d}"
        pending = add_pending_password_reset(contact, password, otp)
        sent, message, visible_otp = send_otp_code(contact, otp, str(user.get("username", "User")), "reset")
        if not sent:
            remove_pending_registration(pending["pending_id"])
            return self.send_json({"error": message}, HTTPStatus.BAD_REQUEST)

        payload = {
            "ok": True,
            "pending_id": pending["pending_id"],
            "message": f"Reset code sent to {contact}.",
        }
        if visible_otp:
            payload["otp"] = visible_otp
            payload["message"] = f"Reset code ready for {contact}."
        return self.send_json(payload)

    def reset_verify(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        pending_id = form.getfirst("pending_id", "").strip()
        otp = form.getfirst("otp", "").strip()
        ok, message, contact = finalize_pending_password_reset(pending_id, otp)
        if not ok or not contact:
            return self.send_json({"error": message}, HTTPStatus.BAD_REQUEST)

        user = find_user_by_contact(contact) or find_user(contact)
        if not user:
            return self.send_json({"error": "Password updated, but the session could not be created."}, HTTPStatus.BAD_REQUEST)

        token = self.session_token(str(user.get("username", "")).strip())
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json")
        self.send_header("Set-Cookie", f"jaffagpt_session={token}; HttpOnly; Path=/; SameSite=Lax")
        body = json.dumps({"ok": True, "username": user.get("username", "")}).encode("utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def google_login(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        credential = form.getfirst("credential", "").strip()
        if not credential:
            return self.send_json({"error": "Missing Google credential."}, HTTPStatus.BAD_REQUEST)

        ok, payload_or_message = verify_google_id_token(credential)
        if not ok:
            return self.send_json({"error": str(payload_or_message)}, HTTPStatus.UNAUTHORIZED)

        payload = payload_or_message
        email = str(payload.get("email", "")).strip()
        display_name = str(payload.get("name", "")).strip()
        user = find_user_by_contact(email)

        if user:
            username = str(user.get("username", "")).strip()
        else:
            username = unique_username(display_name or email.split("@", 1)[0] or "Google User")
            generated_password = uuid.uuid4().hex + uuid.uuid4().hex
            created, message = create_user(username, email, generated_password)
            if not created:
                return self.send_json({"error": message}, HTTPStatus.BAD_REQUEST)
            users = load_users()
            for index, existing in enumerate(users):
                if str(existing.get("username", "")).strip() == username:
                    users[index]["provider"] = "google"
                    save_users(users)
                    break

        token = self.session_token(username)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json")
        self.send_header("Set-Cookie", f"jaffagpt_session={token}; HttpOnly; Path=/; SameSite=Lax")
        body = json.dumps({"ok": True, "username": username, "email": email}).encode("utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def logout(self) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json")
        self.send_header("Set-Cookie", "jaffagpt_session=; HttpOnly; Path=/; Max-Age=0; SameSite=Lax")
        body = json.dumps({"ok": True}).encode("utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_static(self, relative_path: str) -> None:
        path = (STATIC_DIR / relative_path).resolve()
        try:
            path.relative_to(STATIC_DIR.resolve())
        except ValueError:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = "text/html"
        if path.suffix == ".css":
            content_type = "text/css"
        elif path.suffix == ".js":
            content_type = "application/javascript"
        elif path.suffix == ".pdf":
            content_type = "application/pdf"
        elif path.suffix in {".jpg", ".jpeg"}:
            content_type = "image/jpeg"
        elif path.suffix == ".png":
            content_type = "image/png"
        elif path.suffix == ".gif":
            content_type = "image/gif"
        elif path.suffix == ".svg":
            content_type = "image/svg+xml"

        body = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def upload_document(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            return self.send_json({"error": "No file uploaded"}, HTTPStatus.BAD_REQUEST)

        content = file_item.file.read()
        category = form.getfirst("category", "public document").strip() or "public document"
        language = form.getfirst("language", "English").strip() or "English"
        doc_id = str(uuid.uuid4())
        safe_name = re.sub(r"[^a-zA-Z0-9_. -]", "_", file_item.filename)
        if len(content) > MAX_UPLOAD_BYTES:
            return self.send_json(
                {"error": "File is too large. Upload a file smaller than 50 MB."},
                HTTPStatus.BAD_REQUEST,
            )

        try:
            text = extract_text(safe_name, content, language)
        except (RuntimeError, ValueError) as exc:
            text = (
                f"{safe_name} was uploaded and kept in the local document store, "
                f"but readable text could not be extracted automatically. Reason: {exc}"
            )

        (UPLOAD_DIR / f"{doc_id}-{safe_name}").write_bytes(content)

        doc = add_document(safe_name, category, text, doc_id)
        self.send_json({"ok": True, "id": doc.id, "name": safe_name, "characters": len(text)})

    def scan_image(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            return self.send_json({"error": "No image uploaded"}, HTTPStatus.BAD_REQUEST)

        content = file_item.file.read()
        safe_name = re.sub(r"[^a-zA-Z0-9_. -]", "_", file_item.filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix not in ALLOWED_IMAGE_SUFFIXES:
            return self.send_json(
                {"error": "Upload a JPG, PNG, WEBP, or GIF image."},
                HTTPStatus.BAD_REQUEST,
            )

        if len(content) > 20 * 1024 * 1024:
            return self.send_json(
                {"error": "Image is too large. Upload an image smaller than 20 MB."},
                HTTPStatus.BAD_REQUEST,
            )

        language = form.getfirst("language", "English").strip() or "English"
        provider = form.getfirst("provider", "").strip().lower()
        if not provider:
            provider = "openai" if openai_available() else "gemini"
        
        if provider == "gemini":
            answer = call_gemini_vision(safe_name, content, language)
        else:
            answer = call_openai_vision(safe_name, content, language)
            
        text = answer
        if not is_scan_result_usable(answer):
            try:
                ocr_text = local_or_free_ocr(content, language)
                text = ocr_text
                answer = ocr_text
            except Exception as exc:
                text = (
                    f"{safe_name} was uploaded and kept in the local document store, "
                    f"but image text extraction failed. {provider.capitalize()} vision error: {answer}. OCR fallback error: {exc}"
                )
                answer = f"Image uploaded, but OCR failed. Reason: {exc}"
        doc = add_document(f"scan-{safe_name}.txt", "scanned image", text)
        doc_id = doc.id
        (UPLOAD_DIR / f"{doc_id}-{safe_name}").write_bytes(content)
        self.send_json({"ok": True, "id": doc_id, "name": safe_name, "answer": answer})


    def image_to_pdf(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            return self.send_json({"error": "No image uploaded"}, HTTPStatus.BAD_REQUEST)

        content = file_item.file.read()
        safe_name = re.sub(r"[^a-zA-Z0-9_. -]", "_", file_item.filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix not in ALLOWED_IMAGE_SUFFIXES:
            return self.send_json(
                {"error": "Upload a JPG, PNG, WEBP, or GIF image."},
                HTTPStatus.BAD_REQUEST,
            )

        try:
            pdf_bytes = image_to_pdf_bytes(content)
        except (RuntimeError, ValueError) as exc:
            return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        pdf_name = f"{Path(safe_name).stem}.pdf"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Disposition", f'attachment; filename="{pdf_name}"')
        self.send_header("Content-Length", str(len(pdf_bytes)))
        self.end_headers()
        self.wfile.write(pdf_bytes)

    def delete_document(self, doc_id: str) -> None:
        if not re.fullmatch(r"[a-fA-F0-9-]{36}", doc_id):
            return self.send_json({"error": "Invalid document id"}, HTTPStatus.BAD_REQUEST)

        documents = load_documents()
        remaining = [doc for doc in documents if doc.id != doc_id]
        if len(remaining) == len(documents):
            return self.send_json({"error": "Document not found"}, HTTPStatus.NOT_FOUND)

        save_documents(remaining)
        remove_uploaded_files(doc_id)
        self.send_json({"ok": True, "id": doc_id})

    def stream_chat(self, query: str) -> None:
        params = parse_qs(query)
        question = params.get("q", [""])[0].strip()
        mode = params.get("mode", ["simple"])[0]
        language = params.get("language", ["English"])[0]
        task = infer_task(question, params.get("task", [None])[0])
        provider = params.get("provider", [""])[0].strip().lower()
        history = parse_conversation_history(params.get("history", [""])[0])
        
        if not provider:
            provider = choose_default_provider(task, question, history)
                
        try:
            if not question:
                self.send_error(HTTPStatus.BAD_REQUEST)
                return

            documents = load_documents()
            sources = find_sources(question, documents)
            current_datetime = answer_current_datetime(question)
            image = None

            if is_ats_score_request(question):
                resume_doc = select_resume_document(question, documents, sources)
                if not resume_doc:
                    answer = (
                        "Upload your resume document first, then ask for the ATS score again. "
                        "I will score the uploaded resume and suggest specific fixes."
                    )
                else:
                    sources = [
                        {
                            "document": resume_doc.name,
                            "category": resume_doc.category,
                            "excerpt": resume_doc.text[:900],
                            "score": 999,
                        }
                    ]
                    answer = build_ats_score_answer(provider, question, language, resume_doc)
            elif is_song_request(question):
                sources = []
                answer = build_song_response(language)
            elif is_image_to_pdf_request(question):
                sources = []
                uploaded_images = find_uploaded_images()
                if not uploaded_images:
                    answer = (
                        "It looks like you want to convert an image to PDF. "
                        "However, no images have been uploaded yet.\n\n"
                        "Please upload an image first using the **+** button in the composer, "
                        "or use **Image to PDF** in the **Knowledge Base** upload panel."
                    )
                else:
                    img_name, img_path = uploaded_images[0]
                    try:
                        img_bytes = img_path.read_bytes()
                        pdf_bytes = image_to_pdf_bytes(img_bytes)
                        pdf_filename = f"converted_{uuid.uuid4().hex[:8]}_{Path(img_name).stem}.pdf"
                        (STATIC_DIR / pdf_filename).write_bytes(pdf_bytes)
                        pdf_url = f"/static/{pdf_filename}"
                        answer = (
                            f"I found your recently uploaded image **{img_name}** and successfully converted it to PDF!\n\n"
                            f"👉 [**Download Converted PDF**]({pdf_url})\n\n"
                            "Click the link above to save the PDF file to your device."
                        )
                    except Exception as exc:
                        answer = f"I tried to convert your uploaded image **{img_name}** to PDF, but encountered an error: {exc}"
            elif is_image_edit_request(question):
                sources = []
                uploaded_images = find_uploaded_images()
                if not uploaded_images:
                    answer = "Upload an image first, then tell me the edit you want."
                else:
                    img_name, img_path = uploaded_images[0]
                    try:
                        img_bytes = img_path.read_bytes()
                        image = call_openai_image_edit(img_name, img_bytes, question)
                    except OSError as exc:
                        image = {"error": f"I found the uploaded image, but could not open it. Details: {exc}"}

                    if image and image.get("error"):
                        answer = image["error"]
                        image = None
                    elif image:
                        answer = (
                            "Done. I edited the uploaded image with your requested change.\n\n"
                            f"Edit: {question}"
                        )
                    else:
                        answer = "I tried to edit the uploaded image, but the AI service did not return a result. Please try again."
            elif task == "chat" and is_visual_image_question(question) and find_uploaded_images():
                sources = []
                answer = answer_uploaded_image_question(provider, question, language)
                if answer is None:
                    answer = build_ai_answer(provider, question, task, mode, language, [], history)
                    if answer is None:
                        answer = build_answer(question, mode, language, [])
                elif is_image_non_answer(answer):
                    general_answer = build_ai_answer(provider, question, task, mode, language, [], history)
                    if general_answer:
                        answer = general_answer
            elif task == "image":
                sources = []
                image = call_openai_image(question)
                if image and image.get("error"):
                    answer = image["error"]
                    image = None
                elif image:
                    answer = (
                        "Done. I generated the image from your prompt below.\n\n"
                        f"Prompt: {question}"
                    )
                else:
                    answer = build_image_prompt(question, language)
                    image = {"url": f"/api/image?prompt={quote(question)}", "prompt": question}
            elif task == "chat" and current_datetime:
                sources = []
                answer = current_datetime
            elif task == "chat":
                js_snippet = normalize_javascript_snippet(question)
                if js_snippet is not None:
                    sources = []
                    answer = execute_javascript_snippet(js_snippet)
                else:
                    answer = build_ai_answer(provider, question, task, mode, language, sources, history)
                    if answer is None:
                        answer = build_answer(question, mode, language, sources)
            else:
                if task == "chat" and not documents and is_document_lookup(question):
                    answer = build_answer(question, mode, language, sources)
                else:
                    answer = build_ai_answer(provider, question, task, mode, language, sources, history)
                if task == "translate":
                    sources = []
                    if answer is None:
                        answer = build_translation(question, language)
                elif answer is None:
                    answer = build_answer(question, mode, language, sources)

            if answer is None:
                answer = build_answer(question, mode, language, sources)

            if task == "translate":
                sources = []
            elif task == "image":
                sources = []

            # Check if the user asked to generate/download a PDF of notes, sheets, guides
            is_pdf_export_request = "pdf" in question.lower() and any(k in question.lower() for k in ["notes", "write", "generate", "create", "download", "make", "get", "tutorial", "guide", "cheat", "sheet", "plan", "story", "recipe"])

            if is_pdf_export_request and answer and not answer.startswith("⚠️") and not answer.startswith("OpenAI rejected") and not answer.startswith("Gemini connection"):
                try:
                    # Deduce PDF Title
                    title_suggestion = "Generated PDF Notes"
                    if "html" in question.lower():
                        title_suggestion = "HTML Learning Notes"
                    elif "python" in question.lower():
                        title_suggestion = "Python Programming Notes"
                    elif "css" in question.lower():
                        title_suggestion = "CSS Styling Notes"

                    pdf_bytes = create_pdf_from_text(title_suggestion, answer)
                    pdf_filename = f"jaffagpt_{uuid.uuid4().hex[:8]}.pdf"
                    (STATIC_DIR / pdf_filename).write_bytes(pdf_bytes)
                    pdf_url = f"/static/{pdf_filename}"

                    # Append a clean download link
                    answer += (
                        f"\n\n---\n"
                        f"### 📥 PDF Notes Download Available\n"
                        f"I have compiled these notes into a clean PDF document for you:\n"
                        f"👉 [**Download PDF Notes**]({pdf_url})\n\n"
                        f"*(Click the link above to view or save the PDF file directly to your device.)*"
                    )
                except Exception as exc:
                    answer += f"\n\n*(Note: Tried to compile a PDF download, but encountered an error: {exc})*"
        except Exception as exc:
            import traceback
            traceback.print_exc()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/event-stream")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            error_answer = f"Sorry, the server hit an unexpected error while preparing this answer: {exc}"
            for chunk in sse_chunks(error_answer, [], None):
                self.wfile.write(chunk)
                self.wfile.flush()
            return

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        try:
            for chunk in sse_chunks(answer, sources, image):
                self.wfile.write(chunk)
                self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError):
            return


def main() -> None:
    ensure_dirs()
    preferred_port = int(os.environ.get("PORT") or (sys.argv[1] if len(sys.argv) > 1 else 8000))
    server = None
    port = preferred_port
    for candidate in range(preferred_port, preferred_port + 20):
        try:
            server = ThreadingHTTPServer(("127.0.0.1", candidate), PublicAIHandler)
            port = candidate
            break
        except OSError:
            continue

    if server is None:
        raise SystemExit(f"No available local port from {preferred_port} to {preferred_port + 19}.")

    print(f"JaffaGPT running at http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
